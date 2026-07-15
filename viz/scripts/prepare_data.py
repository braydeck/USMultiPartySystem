#!/usr/bin/env python3
"""Prepare CSV data from simulation outputs into JSON files for the React viz app."""

import csv
import json
import os
from pathlib import Path
from collections import defaultdict

OUTPUTS          = Path(__file__).parent.parent.parent / "data" / "outputs"
FD_DIR           = OUTPUTS / "factor_deviation"
PURE_MULTI_DIR   = OUTPUTS / "pure_multi"
# Parallel run with Solidarity (cluster 2) dissolved — produced by NO_STY=1 pipeline.
PURE_MULTI_NOSTY_DIR = OUTPUTS / "pure_multi_nosty"
# 'Current participation' runs — ballots weighted by each cluster's validated 2024
# turnout (TURNOUT_WEIGHT=1). Two coordination variants: all-parties and no-Solidarity.
PURE_MULTI_TURNOUT_DIR       = OUTPUTS / "pure_multi_turnout"
PURE_MULTI_NOSTY_TURNOUT_DIR = OUTPUTS / "pure_multi_nosty_turnout"
# Parallel 10-party run (C7/WFP activated) — produced by INCLUDE_C7=1 pipeline.
PURE_MULTI_C7_DIR = OUTPUTS / "pure_multi_c7"
FD_TRIPLE_DIR         = OUTPUTS / "factor_deviation_triple"
PURE_MULTI_TRIPLE_DIR = OUTPUTS / "pure_multi_triple"
RESULTS          = Path(__file__).parent.parent.parent / "results"
DATA_OUT = Path(__file__).parent.parent / "src" / "data"
DATA_OUT.mkdir(parents=True, exist_ok=True)

CLUSTER_TO_PARTY = {
    "0": "CON", "1": "LBR", "2": "STY", "3": "NAT",
    "4": "LIB", "5": "POP", "6": "CUP", "7": "OAO", "8": "DSA", "9": "PRG",
}

# Soft-weighted national population shares per cluster (excluding C7/Blue Dogs, renormalized to 100%)
# Derived from: np.average(prob_cluster_k, weights=commonpostweight), then renormalized.
# Matches the GMM posterior scoring used in ballot generation.
NATIONAL_POP_SHARES = {
    0: 18.73,   # CON
    1: 15.63,   # SD
    2: 15.04,   # STY
    3:  9.19,   # NAT
    4:  9.30,   # LIB
    5: 11.01,   # POP
    6:  9.86,   # CUP
    8:  6.27,   # DSA
    9:  4.96,   # PRG
}


PARTY_NAMES = {
    "CON": "Conservative", "LBR": "Labor", "STY": "Solidarity",
    "NAT": "Nationalist", "LIB": "Liberal", "POP": "Populist",
    "CUP": "Civic Union Party", "OAO": "Order and Opportunity Party",
    "DSA": "Democratic Socialists", "PRG": "Progressive",
}

def read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def write_json(data, name):
    path = DATA_OUT / name
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, separators=(",", ":"))
    print(f"  Wrote {path.name} ({os.path.getsize(path):,} bytes)")




# ── LF senate vote model helpers ─────────────────────────────────────────────

_LF_CLUSTER_MAP: dict = {
    "PRG_dsa": (9, 8), "DSA_prg": (8, 9), "DSA_lib": (8, 4),
    "LIB_dsa": (4, 8), "LIB_sd":  (4, 1), "SD_lib":  (1, 4),
    "SD_sty":  (1, 2), "STY_sd":  (2, 1), "STY_ctr": (2, 6),
    "CUP_sty": (6, 2), "CUP_con": (6, 0), "CON_ctr": (0, 6),
    "CON_ref": (0, 5), "POP_con": (5, 0), "POP_nat": (5, 3),
    "NAT_ref": (3, 5),
}
_PURE_CLUSTER: dict = {
    "CON": 0, "LBR": 1, "STY": 2, "NAT": 3, "LIB": 4,
    "POP": 5, "CUP": 6, "DSA": 8, "PRG": 9,
    "OAO": 7,   # Order and Opportunity Party (cluster 7)
}

def _lf_senator_support(code: str, cluster_row: dict) -> float:
    """Return the % supporting a bill for a given LF senator, using cluster stats row."""
    def _c(k: int) -> float:
        return float(cluster_row.get(f"c{k}") or 0)
    if code in _PURE_CLUSTER:
        return _c(_PURE_CLUSTER[code])
    if code in _LF_CLUSTER_MAP:
        p, s = _LF_CLUSTER_MAP[code]
        return 0.80 * _c(p) + 0.20 * _c(s)
    return 0.0

def _fd_prob_pass(seat_counts: dict, fd_by_var: dict, majority: int = 26) -> dict:
    """Compute prob_pass/verdict for FD senate using Normal approximation.
    seat_counts: {senator_code: n_seats}
    fd_by_var: {variable: fd_stats_row}  — % Supporting rows for FD candidates
    Returns: {variable: {"prob_pass": float, "verdict": str}}
    """
    import math
    results = {}
    for var, row in fd_by_var.items():
        mu = 0.0
        var_sum = 0.0
        for code, n in seat_counts.items():
            p = float(row.get(code) or 0) / 100.0
            mu += n * p
            var_sum += n * p * (1 - p)
        sigma = math.sqrt(var_sum) if var_sum > 0 else 1e-6
        z = (majority - 0.5 - mu) / sigma
        def _ncdf(x: float) -> float:
            if x > 6: return 1.0
            if x < -6: return 0.0
            t = 1.0 / (1.0 + 0.2316419 * abs(x))
            d = 0.3989422803 * math.exp(-0.5 * x * x)
            p_approx = d * t * (0.3193815 + t * (-0.3565638 + t * (1.7814779 + t * (-1.8212560 + t * 1.3302744))))
            return p_approx if x < 0 else 1.0 - p_approx
        prob = 1.0 - _ncdf(z)
        verdict = "PASS" if prob >= 0.7 else "FAIL" if prob <= 0.3 else "TOSS-UP"
        results[var] = {"prob_pass": round(prob, 4), "verdict": verdict}
    return results


def _lf_prob_pass(seat_counts: dict, cluster_by_var: dict, majority: int = 26) -> tuple:
    """Compute prob_pass and verdict for LF senate using Normal approximation.
    seat_counts: {senator_code: n_seats}
    cluster_by_var: {variable: cluster_row_dict}
    Returns: {variable: {"prob_pass": float, "verdict": str}}
    """
    import math
    results = {}
    for var, crow in cluster_by_var.items():
        mu = 0.0
        var_sum = 0.0
        for code, n in seat_counts.items():
            p = _lf_senator_support(code, crow) / 100.0
            mu += n * p
            var_sum += n * p * (1 - p)
        sigma = math.sqrt(var_sum) if var_sum > 0 else 1e-6
        # Continuity-corrected P(Y >= majority)
        z = (majority - 0.5 - mu) / sigma
        # Normal CDF approximation (Abramowitz & Stegun 26.2.17)
        def _ncdf(x: float) -> float:
            if x > 6:
                return 1.0
            if x < -6:
                return 0.0
            t = 1.0 / (1.0 + 0.2316419 * abs(x))
            d = 0.3989422803 * math.exp(-0.5 * x * x)
            p_approx = d * t * (0.3193815 + t * (-0.3565638 + t * (1.7814779 + t * (-1.8212560 + t * 1.3302744))))
            return p_approx if x < 0 else 1.0 - p_approx
        prob = 1.0 - _ncdf(z)
        if prob >= 0.7:
            verdict = "PASS"
        elif prob <= 0.3:
            verdict = "FAIL"
        else:
            verdict = "TOSS-UP"
        results[var] = {"prob_pass": round(prob, 4), "verdict": verdict}
    return results


# ---------- senateVoteModel.json ----------
def build_senate_vote_model(rm_dir=PURE_MULTI_DIR, out_name="senateVoteModel.json"):
    rows = read_csv(RESULTS / "vote_model.csv")

    # Load cluster stats for vote model computation
    cluster_rows = read_csv(OUTPUTS / "profiles" / "cluster_stats.csv")
    cluster_by_var = {
        r["variable"]: r
        for r in cluster_rows
        if r.get("stat_label") == "% Supporting" and r.get("type") == "binary"
    }

    # ── Factor Deviation senate vote model ──────────────────────────────────
    fd_stats_rows = read_csv(FD_DIR / "profiles" / "factor_deviation_stats.csv")
    fd_by_var = {
        r["variable"]: r
        for r in fd_stats_rows
        if r.get("stat_label") == "% Supporting" and r.get("type") == "binary"
           and r.get("variable", "").startswith("CC24_")
    }

    cond_fd_seats = {}
    irv_fd_seats = {}
    for r in read_csv(FD_DIR / "senate" / "senate_composition.csv"):
        cond_fd_seats[r["senator_code"]] = cond_fd_seats.get(r["senator_code"], 0) + 1
    for r in read_csv(FD_DIR / "senate" / "senate_irv_composition.csv"):
        irv_fd_seats[r["senator_code"]] = irv_fd_seats.get(r["senator_code"], 0) + 1

    fd_cond_results = _fd_prob_pass(cond_fd_seats, fd_by_var)
    fd_irv_results  = _fd_prob_pass(irv_fd_seats,  fd_by_var)

    # FD presidential winners — read dynamically from simulation outputs
    fd_irv_winner = None
    for r in read_csv(FD_DIR / "irv" / "irv_presidential_national_2028.csv"):
        if r.get("winner") == "True":
            fd_irv_winner = r["candidate_code"]
    fd_cond_winner = None
    _cm_rows = list(read_csv(FD_DIR / "irv" / "condorcet_matchups_2028.csv"))
    if _cm_rows:
        fd_cond_winner = _cm_rows[0].get("condorcet_winner") or None

    fd_pres_irv_signs = {}
    fd_pres_irv_pct   = {}
    fd_pres_cond_signs = {}
    fd_pres_cond_pct   = {}
    for var, row in fd_by_var.items():
        irv_sup  = float(row.get(fd_irv_winner)  or 0) if fd_irv_winner  else 0.0
        cond_sup = float(row.get(fd_cond_winner) or 0) if fd_cond_winner else 0.0
        fd_pres_irv_signs[var]   = "SIGN" if irv_sup  > 50 else "VETO"
        fd_pres_irv_pct[var]     = round(irv_sup, 2)
        fd_pres_cond_signs[var]  = "SIGN" if cond_sup > 50 else "VETO"
        fd_pres_cond_pct[var]    = round(cond_sup, 2)
    # Alias for backwards compat — IRV winner is the "default" FD president
    fd_pres_signs = fd_pres_irv_signs
    fd_pres_pct   = fd_pres_irv_pct

    # ── Raw Multi presidential winners ───────────────────────────────────────
    rm_irv_winner = None
    for r in read_csv(rm_dir / "irv" / "irv_presidential_national_2028.csv"):
        if r.get("winner", "").strip() == "True":
            rm_irv_winner = r["candidate_code"]   # e.g. "SD_1"
    rm_cond_winner = None
    _rm_cm = list(read_csv(rm_dir / "irv" / "condorcet_matchups_2028.csv"))
    if _rm_cm:
        rm_cond_winner = _rm_cm[0].get("condorcet_winner") or None   # e.g. "CUP_1"

    def _rm_party(code):
        return code.rsplit("_", 1)[0] if code else ""

    presRawMultiIRV_signs,  presRawMultiIRV_pct  = {}, {}
    presRawMultiCond_signs, presRawMultiCond_pct = {}, {}
    for var, crow in cluster_by_var.items():
        irv_party  = _rm_party(rm_irv_winner)
        cond_party = _rm_party(rm_cond_winner)
        irv_sup  = _lf_senator_support(irv_party,  crow)
        cond_sup = _lf_senator_support(cond_party, crow)
        presRawMultiIRV_signs[var]  = "SIGN" if irv_sup  > 50 else "VETO"
        presRawMultiIRV_pct[var]    = round(irv_sup, 2)
        presRawMultiCond_signs[var] = "SIGN" if cond_sup > 50 else "VETO"
        presRawMultiCond_pct[var]   = round(cond_sup, 2)

    # Raw Multi senate seat counts (strip _N suffix to aggregate by pure party)
    cond_rm_seats, irv_rm_seats = {}, {}
    for r in read_csv(rm_dir / "senate" / "senate_composition.csv"):
        party = r["senator_code"].rsplit("_", 1)[0]
        cond_rm_seats[party] = cond_rm_seats.get(party, 0) + 1
    for r in read_csv(rm_dir / "senate" / "senate_irv_composition.csv"):
        party = r["senator_code"].rsplit("_", 1)[0]
        irv_rm_seats[party] = irv_rm_seats.get(party, 0) + 1

    rm_cond_results = _lf_prob_pass(cond_rm_seats, cluster_by_var)
    rm_irv_results  = _lf_prob_pass(irv_rm_seats,  cluster_by_var)

    # Corrected pure presidential signing — actual winner is SD, not STY as in legacy vote_model.csv
    presPure_signs, presPure_pct = {}, {}
    for var, crow in cluster_by_var.items():
        sd_sup = _lf_senator_support("LBR", crow)
        presPure_signs[var] = "SIGN" if sd_sup > 50 else "VETO"
        presPure_pct[var]   = round(sd_sup, 2)

    # State STV house (873 seats, majority = 437) — for LegislationTab house column
    stv_house_seat_counts = {}
    stv_house_total = 0
    for r in read_csv(OUTPUTS / "house" / "house_seat_summary.csv"):
        stv_house_seat_counts[r["party"]] = int(r["seats"])
        stv_house_total += int(r["seats"])
    stv_house_majority = stv_house_total // 2 + 1
    stv_house_results = _lf_prob_pass(stv_house_seat_counts, cluster_by_var, majority=stv_house_majority)

    # Compute SD/CON (Condorcet blended president) signing decisions from chamber profile
    senate_prof_rows = read_csv(OUTPUTS / "senate" / "senate_chamber_profile.csv")
    sdcon_pct = {}
    for r in senate_prof_rows:
        if r.get("stat_label") == "% Supporting" and r.get("variable", "").startswith("CC24_"):
            try:
                sdcon_pct[r["variable"]] = float(r["SD/CON"])
            except (KeyError, ValueError):
                pass

    out = []
    for r in rows:
        var = r["variable"]
        sdcon_support = sdcon_pct.get(var)
        pres_mixed_cond_pct = round(sdcon_support, 2) if sdcon_support is not None else float(r.get("pres_mixed_support_pct", 0))
        pres_mixed_cond_signs = ("SIGN" if sdcon_support > 50 else "VETO") if sdcon_support is not None else r["pres_mixed_signs"]

        out.append({
            "variable": var,
            "domain": r["domain"],
            "question": r["question"],
            "overallPct": float(r["overall_pct"]),
            # Mixed senate scenarios (new keys + legacy aliases for UnifiedBillTable)
            "condMixedProbPass": float(r["senate_cond_prob_pass"]),
            "condMixedVerdict": r["senate_cond_verdict"],
            "irvMixedProbPass": float(r["senate_irv_prob_pass"]),
            "irvMixedVerdict": r["senate_irv_verdict"],
            # Legacy aliases (UnifiedBillTable reads these)
            "condProbPass": float(r["senate_cond_prob_pass"]),
            "condVerdict": r["senate_cond_verdict"],
            "irvProbPass": float(r["senate_irv_prob_pass"]),
            "irvVerdict": r["senate_irv_verdict"],
            # Pure senate scenarios
            "condPureProbPass": float(r["senate_cond_pure_prob_pass"]),
            "condPureVerdict": r["senate_cond_pure_verdict"],
            "irvPureProbPass": float(r["senate_irv_pure_prob_pass"]),
            "irvPureVerdict": r["senate_irv_pure_verdict"],
            # Factor Deviation senate scenarios
            "condFDProbPass": fd_cond_results.get(var, {}).get("prob_pass", 0.0),
            "condFDVerdict":  fd_cond_results.get(var, {}).get("verdict", "N/A"),
            "irvFDProbPass":  fd_irv_results.get(var, {}).get("prob_pass", 0.0),
            "irvFDVerdict":   fd_irv_results.get(var, {}).get("verdict", "N/A"),
            # FD president — separate fields for IRV and Condorcet winners
            "presFDSigns":     fd_pres_irv_signs.get(var, "VETO"),   # alias = IRV
            "presFDPct":       fd_pres_irv_pct.get(var, 0.0),
            "presFDIRVSigns":  fd_pres_irv_signs.get(var, "VETO"),
            "presFDIRVPct":    fd_pres_irv_pct.get(var, 0.0),
            "presFDCondSigns": fd_pres_cond_signs.get(var, "VETO"),
            "presFDCondPct":   fd_pres_cond_pct.get(var, 0.0),
            # Presidential sign + support pct
            "presMixedSigns": r["pres_mixed_signs"],         # CON/SD (blended IRV winner)
            "presMixedPct":   float(r.get("pres_mixed_support_pct", 0)),
            "presMixedCondSigns": pres_mixed_cond_signs,     # SD/CON (blended Condorcet)
            "presMixedCondPct": pres_mixed_cond_pct,
            "presPureSigns": presPure_signs.get(var, "VETO"),  # SD (actual pure winner)
            "presPurePct":   presPure_pct.get(var, 0.0),
            # Raw Multi senate scenarios
            "condRawMultiProbPass": rm_cond_results.get(var, {}).get("prob_pass", 0.0),
            "condRawMultiVerdict":  rm_cond_results.get(var, {}).get("verdict", "N/A"),
            "irvRawMultiProbPass":  rm_irv_results.get(var, {}).get("prob_pass", 0.0),
            "irvRawMultiVerdict":   rm_irv_results.get(var, {}).get("verdict", "N/A"),
            # Raw Multi president (SD_1 IRV, CUP_1 Condorcet)
            "presRawMultiIRVSigns":  presRawMultiIRV_signs.get(var, "VETO"),
            "presRawMultiIRVPct":    presRawMultiIRV_pct.get(var, 0.0),
            "presRawMultiCondSigns": presRawMultiCond_signs.get(var, "VETO"),
            "presRawMultiCondPct":   presRawMultiCond_pct.get(var, 0.0),
            # State STV house (for LegislationTab house column)
            "houseStvProbPass": stv_house_results.get(var, {}).get("prob_pass", 0.0),
            "houseStvVerdict":  stv_house_results.get(var, {}).get("verdict", "N/A"),
        })
    write_json(out, out_name)


# ---------- houseSeats.json ----------
def build_house_seats(src_csv=None, out_name="houseSeats.json",
                      include_c7=True, pop_shares=None):
    if src_csv is None:
        src_csv = OUTPUTS / "pure_multi" / "house" / "stv_seat_summary.csv"
    if pop_shares is None:
        pop_shares = _national_pop_shares_10() if include_c7 else NATIONAL_POP_SHARES
    rows = read_csv(src_csv)
    out = []
    for r in rows:
        cluster = int(r["party"])
        if cluster == 7 and not include_c7:  # skip Blue Dogs (C7) in canonical 9-party
            continue
        out.append({
            "party": cluster,
            "partyName": r["party_name"],
            "urban": int(r["URBAN"]),
            "suburban": int(r["SUBURBAN"]),
            "rural": int(r["RURAL"]),
            "national": int(r["NATIONAL"]),
            "pctNational": float(r["pct_national"]),
            "pctPopulation": pop_shares.get(cluster, 0.0),
        })
    write_json(out, out_name)


def build_house_transfers():
    """Compute transfer preference matrix from party ballots."""
    ballot_path = PURE_MULTI_DIR / "party_ballots.csv"
    if not ballot_path.exists():
        ballot_path = PURE_MULTI_DIR / "presidential_ballots.csv"
    if not ballot_path.exists():
        write_json([], "houseTransfers.json")
        return

    # Read ballots and weights using csv reader
    ballot_rows = read_csv(str(ballot_path))
    weight_rows = read_csv(str(Path(__file__).parent.parent.parent / "data" / "processed" / "efa_factor_scores.csv"))

    def base_party(code):
        if "_" in code and code.rsplit("_", 1)[1].isdigit():
            return code.rsplit("_", 1)[0]
        return code

    # Discover base party codes
    base_codes = sorted(set(base_party(r["rank_1"]) for r in ballot_rows))

    transfers = {src: {dst: 0.0 for dst in base_codes} for src in base_codes}
    first_totals = {c: 0.0 for c in base_codes}

    for i, br in enumerate(ballot_rows):
        w = float(weight_rows[i]["commonpostweight"])
        first  = base_party(br["rank_1"])
        second = base_party(br["rank_2"])
        first_totals[first] += w
        transfers[first][second] += w

    out = []
    for src in base_codes:
        total = first_totals[src] or 1.0
        dests = []
        for dst in base_codes:
            if dst == src:
                continue
            pct = transfers[src][dst] / total * 100
            if pct > 0.5:
                dests.append({"party": dst, "pct": round(pct, 1)})
        dests.sort(key=lambda x: -x["pct"])
        out.append({
            "source": src,
            "totalVoters": round(first_totals[src], 0),
            "destinations": dests,
        })
    write_json(out, "houseTransfers.json")


def build_fd_variant_attraction():
    """For each FD variant, show where first-choice voters come from (home vs cross-party)."""
    ballot_path = FD_DIR / "ballots.csv"
    if not ballot_path.exists():
        write_json([], "fdVariantAttraction.json")
        return

    ballot_rows = read_csv(str(ballot_path))
    typology_rows = read_csv(str(Path(__file__).parent.parent.parent / "data" / "processed" / "typology_cluster_assignments.csv"))

    CLUSTER_TO_PARTY = {"0":"CON","1":"LBR","2":"STY","3":"NAT","4":"LIB","5":"POP","6":"CUP","7":"OAO","8":"DSA","9":"PRG"}

    # variant → {source_party: weighted_count}
    variant_sources: dict = {}
    variant_totals: dict = {}

    for i, br in enumerate(ballot_rows):
        variant = br.get("rank_1", "")
        cluster = typology_rows[i].get("cluster", "")
        home = CLUSTER_TO_PARTY.get(str(cluster), "")
        if home == "OAO" or not home:
            continue
        w = float(typology_rows[i].get("commonpostweight", 1))
        if variant not in variant_sources:
            variant_sources[variant] = {}
            variant_totals[variant] = 0.0
        variant_sources[variant][home] = variant_sources[variant].get(home, 0.0) + w
        variant_totals[variant] += w

    out = []
    for variant in sorted(variant_sources.keys()):
        total = variant_totals[variant]
        if total < 50:
            continue
        parts = variant.split("_")
        party = parts[0]
        if len(parts) == 3:
            direction = parts[1]   # hi / lo
            axis = parts[2]        # so / ae / pc / rt
        else:
            axis = "base"
            direction = "base"

        home_pct = variant_sources[variant].get(party, 0) / total * 100
        sources = []
        for src_party, w in sorted(variant_sources[variant].items(), key=lambda x: -x[1]):
            pct = w / total * 100
            if pct >= 2:
                sources.append({"party": src_party, "pct": round(pct, 1)})

        out.append({
            "variant": variant,
            "party": party,
            "axis": axis,
            "direction": direction,
            "totalVoters": round(total, 0),
            "homePct": round(home_pct, 1),
            "crossPct": round(100 - home_pct, 1),
            "sources": sources,
        })

    out.sort(key=lambda x: -x["crossPct"])
    write_json(out, "fdVariantAttraction.json")


def build_house_seats_gauss():
    """Gaussian reference run — reads the _gauss suffix file."""
    rows = read_csv(OUTPUTS / "pure_multi" / "house" / "stv_seat_summary_gauss.csv")
    pop_shares = _national_pop_shares_10()
    out = []
    for r in rows:
        cluster = int(r["party"])
        out.append({
            "party": cluster,
            "partyName": r["party_name"],
            "urban": int(r["URBAN"]),
            "suburban": int(r["SUBURBAN"]),
            "rural": int(r["RURAL"]),
            "national": int(r["NATIONAL"]),
            "pctNational": float(r["pct_national"]),
            "pctPopulation": pop_shares.get(cluster, 0.0),
        })
    write_json(out, "houseSeatsProbBased.json")


# ---------- houseVoteModel.json ----------
def build_house_vote_model(rm_dir=PURE_MULTI_DIR, out_name="houseVoteModel.json",
                           include_c7=True):
    rows = read_csv(OUTPUTS / "house_vote_model.csv")

    # Load cluster stats for State STV house probability computation
    cluster_rows = read_csv(OUTPUTS / "profiles" / "cluster_stats.csv")
    cluster_by_var_h = {
        r["variable"]: r
        for r in cluster_rows
        if r.get("stat_label") == "% Supporting" and r.get("type") == "binary"
    }

    # State STV seat counts (from new house pipeline)
    stv_seat_counts = {}
    total_stv = 0
    for r in read_csv(OUTPUTS / "house" / "house_seat_summary.csv"):
        stv_seat_counts[r["party"]] = int(r["seats"])
        total_stv += int(r["seats"])
    stv_majority = total_stv // 2 + 1
    stv_results = _lf_prob_pass(stv_seat_counts, cluster_by_var_h, majority=stv_majority)

    # Raw Multi house seats (pure_multi — integer cluster id -> party code)
    _cluster_to_party = {v: k for k, v in _PURE_CLUSTER.items()}
    rm_house_seats: dict = {}
    rm_house_total = 0
    for r in read_csv(rm_dir / "house" / "stv_seat_summary.csv"):
        cluster = int(r["party"])
        if cluster == 7 and not include_c7:
            continue
        code = _cluster_to_party.get(cluster, str(cluster))
        rm_house_seats[code] = rm_house_seats.get(code, 0) + int(r["NATIONAL"])
        rm_house_total += int(r["NATIONAL"])
    rm_house_majority = rm_house_total // 2 + 1
    rm_house_results = _lf_prob_pass(rm_house_seats, cluster_by_var_h, majority=rm_house_majority)

    # FD house seats (factor_deviation — party column is already a code, group by base party)
    fd_house_seats: dict = {}
    fd_house_total = 0
    for r in read_csv(FD_DIR / "house" / "stv_seat_summary.csv"):
        code = r["party"]
        fd_house_seats[code] = fd_house_seats.get(code, 0) + int(r["NATIONAL"])
        fd_house_total += int(r["NATIONAL"])
    fd_house_majority = fd_house_total // 2 + 1
    fd_house_results = _lf_prob_pass(fd_house_seats, cluster_by_var_h, majority=fd_house_majority)

    # Triple Wyoming — Raw Multi house seats
    rm_triple_seats: dict = {}
    rm_triple_total = 0
    rm_triple_path = PURE_MULTI_TRIPLE_DIR / "house" / "stv_seat_summary.csv"
    if rm_triple_path.exists():
        for r in read_csv(rm_triple_path):
            cluster = int(r["party"])
            if cluster == 7:
                continue
            code = _cluster_to_party.get(cluster, str(cluster))
            rm_triple_seats[code] = rm_triple_seats.get(code, 0) + int(r["NATIONAL"])
            rm_triple_total += int(r["NATIONAL"])
    rm_triple_majority = rm_triple_total // 2 + 1 if rm_triple_total else 1
    rm_triple_results = _lf_prob_pass(rm_triple_seats, cluster_by_var_h, majority=rm_triple_majority) if rm_triple_seats else {}

    # Triple Wyoming — FD house seats
    fd_triple_seats: dict = {}
    fd_triple_total = 0
    fd_triple_path = FD_TRIPLE_DIR / "house" / "stv_seat_summary.csv"
    if fd_triple_path.exists():
        for r in read_csv(fd_triple_path):
            code = r["party"]
            fd_triple_seats[code] = fd_triple_seats.get(code, 0) + int(r["NATIONAL"])
            fd_triple_total += int(r["NATIONAL"])
    fd_triple_majority = fd_triple_total // 2 + 1 if fd_triple_total else 1
    fd_triple_results = _lf_prob_pass(fd_triple_seats, cluster_by_var_h, majority=fd_triple_majority) if fd_triple_seats else {}

    out = []
    for r in rows:
        var = r["variable"]
        stv = stv_results.get(var, {"prob_pass": 0.0, "verdict": "N/A"})
        rm  = rm_house_results.get(var, {"prob_pass": 0.0, "verdict": "N/A"})
        fd  = fd_house_results.get(var, {"prob_pass": 0.0, "verdict": "N/A"})
        rmt = rm_triple_results.get(var, {"prob_pass": 0.0, "verdict": "N/A"})
        fdt = fd_triple_results.get(var, {"prob_pass": 0.0, "verdict": "N/A"})
        out.append({
            "variable": var,
            "domain": r["domain"],
            "question": r["question"],
            "overallPct": float(r["overall_pct"]),
            "probPass": float(r["house_prob_pass"]),
            "verdict": r["house_verdict"],
            "houseStvProbPass":      stv["prob_pass"],
            "houseStvVerdict":       stv["verdict"],
            "houseRawMultiProbPass": rm["prob_pass"],
            "houseRawMultiVerdict":  rm["verdict"],
            "houseFDProbPass":       fd["prob_pass"],
            "houseFDVerdict":        fd["verdict"],
            "houseRawMultiTripleProbPass": rmt["prob_pass"],
            "houseRawMultiTripleVerdict":  rmt["verdict"],
            "houseFDTripleProbPass":       fdt["prob_pass"],
            "houseFDTripleVerdict":        fdt["verdict"],
        })
    write_json(out, out_name)


# ---------- houseStateMap.json ----------
def _compute_state_pop_shares(include_c7: bool = True) -> dict:
    """Compute soft-weighted population shares per state.

    Canonical (include_c7=False): C7 dropped and remaining 9 renormalized to 100%.
    WFP (include_c7=True): C7 kept as WFP; all 10 clusters sum to ~100% (no renorm).
    """
    import numpy as np
    efa_path = Path(__file__).parent.parent.parent / "data" / "processed" / "efa_factor_scores.csv"
    typ_path = Path(__file__).parent.parent.parent / "data" / "processed" / "typology_cluster_assignments.csv"
    efa_rows = read_csv(efa_path)
    typ_rows = read_csv(typ_path)

    PARTY_CODES = {0: "CON", 1: "LBR", 2: "STY", 3: "NAT", 4: "LIB", 5: "POP", 6: "CUP", 8: "DSA", 9: "PRG"}
    if include_c7:
        PARTY_CODES = {**PARTY_CODES, 7: "OAO"}
    result: dict = {}
    states = set()
    for r in efa_rows:
        v = r.get("inputstate", "")
        if v:
            try:
                states.add(int(float(v)))
            except (ValueError, TypeError):
                pass
    for st_int in sorted(states):
        st_str_candidates = [str(st_int), str(float(st_int)), str(st_int) + ".0"]
        idxs = [i for i, r in enumerate(efa_rows) if r.get("inputstate") in st_str_candidates]
        ws = [float(efa_rows[i].get("commonpostweight", 1)) for i in idxs]
        total_w = sum(ws)
        if total_w == 0:
            continue
        shares: dict = {}
        c7_share = 0.0
        for k in range(10):
            col = f"prob_cluster_{k}"
            val = sum(float(typ_rows[i].get(col, 0)) * ws[j] for j, i in enumerate(idxs)) / total_w * 100
            if k == 7 and not include_c7:
                c7_share = val
            elif k in PARTY_CODES:
                shares[PARTY_CODES[k]] = val
        renorm = 100.0 / (100.0 - c7_share) if c7_share < 100 else 1.0
        for code in shares:
            shares[code] = round(shares[code] * renorm, 2)
        fips2 = str(st_int).zfill(2)
        result[fips2] = shares
    return result


def build_house_state_map(src_dir=None, out_name="houseStateMap.json", include_c7=True):
    """Aggregate house STV results by state to find plurality party per state."""
    if src_dir is None:
        src_dir = OUTPUTS / "pure_multi"
    rows = read_csv(src_dir / "house" / "stv_results_by_district.csv")
    pod_path = OUTPUTS / "state_pod_assignments.csv"
    if not pod_path.exists():
        pod_path = src_dir / "state_pod_assignments.csv"
    if not pod_path.exists():
        pod_path = PURE_MULTI_DIR / "state_pod_assignments.csv"
    pod_rows = read_csv(pod_path) if pod_path.exists() else []
    abbr_by_fips = {r["state_fips"].zfill(2): r["state_abbr"] for r in pod_rows}

    state_pop_shares = _compute_state_pop_shares(include_c7=include_c7)

    state_seats = defaultdict(lambda: defaultdict(int))
    for row in rows:
        fips = row["state_fips"].zfill(2)
        for i in range(20):
            v = row.get(f"elected_{i}", "").strip()
            if v:
                state_seats[fips][v] += 1

    out = {}
    for fips, counts in state_seats.items():
        total = sum(counts.values())
        plurality = max(counts, key=counts.get)
        out[fips] = {
            "stateAbbr": abbr_by_fips.get(fips, fips),
            "pluralityParty": plurality,
            "totalSeats": total,
            "seats": dict(counts),
            "popShares": state_pop_shares.get(fips, {}),
        }
    write_json(out, out_name)


# ---------- coalitionProfiles.json ----------
def build_coalition_profiles():
    rows = read_csv(OUTPUTS / "coalitions" / "coalition_type_profiles.csv")
    out = []
    for r in rows:
        if '/' in r["type"]:
            continue  # skip legacy blended senate entries
        out.append({
            "type": r["type"],
            "chamber": r["chamber"],
            "F1": float(r["F1_security_order"]),
            "F2": float(r["F2_electoral_skepticism"]),
            "F3": float(r["F3_government_distrust"]),
            "F4": float(r["F4_religious_traditionalism"]),
            "F5": float(r["F5_populist_conservatism"]),
            "seatsHouse": int(r["seats_house"]),
            "seatsSenateCondorcet": int(r.get("seats_senate_cond", 0)),
            "seatsSenateIRV": int(r.get("seats_senate_irv", 0)),
        })
    write_json(out, "coalitionProfiles.json")


# ---------- transferMatrix.json ----------
def build_transfer_matrix():
    rows = read_csv(OUTPUTS / "No_C7_canonical" / "transfer_matrix_10party.csv")
    parties = [k for k in rows[0].keys() if k != "party_a"]
    matrix = {}
    for row in rows:
        src = row["party_a"]
        if not src:
            continue
        matrix[src] = {p: float(row[p]) for p in parties if row.get(p, "0") != "0"}
    write_json({"parties": parties, "matrix": matrix}, "transferMatrix.json")


# ---------- clusterProfiles.json ----------
def collect_cluster_variables(rows, include_c7=True):
    """Build a variable dict for each cluster covering all binary/likert policy vars + demographics.

    Sources:
      - binary (% Supporting) and binary_agree (% Agreeing): included directly
      - likert5: agree-% computed as % Strongly Agree + % Agree
      - pew_churatd: weekly+ church attendance (% More than once/week + % Once/week)
      - race / gender4 categorical: specific buckets included as demographic facts
    """
    cluster_ids = [str(i) for i in range(10) if include_c7 or str(i) != "7"]
    result = {cid: {} for cid in cluster_ids}

    # Phase 1: binary and binary_agree
    for r in rows:
        typ = r.get("type", "")
        lbl = r.get("stat_label", "")
        if not ((typ == "binary" and lbl == "% Supporting") or
                (typ == "binary_agree" and lbl == "% Agreeing")):
            continue
        var = r["variable"]
        try:
            overall = float(r.get("overall", "") or 0)
        except (ValueError, TypeError):
            continue
        for cid in cluster_ids:
            try:
                val = float(r.get(f"c{cid}", "") or 0)
                result[cid][var] = {
                    "pct": round(val, 1),
                    "question": r.get("question", var),
                    "domain": r.get("domain", ""),
                    "diffPp": round(val - overall, 1),
                    "overall": round(overall, 1),
                }
            except (ValueError, TypeError):
                pass

    # Phase 2: Likert5 agree-% (% Strongly Agree + % Agree)
    likert_meta = {}   # var -> {question, domain}
    likert_sa   = {}   # var -> row for % Strongly Agree
    likert_ag   = {}   # var -> row for % Agree
    for r in rows:
        var = r["variable"]
        typ = r.get("type", "")
        lbl = r.get("stat_label", "")
        if typ == "likert5":
            likert_meta[var] = {"question": r.get("question", var), "domain": r.get("domain", "")}
        elif typ == "likert5_dist":
            if lbl == "% Strongly Agree":
                likert_sa[var] = r
            elif lbl == "% Agree":
                likert_ag[var] = r

    for var, meta in likert_meta.items():
        sa_row = likert_sa.get(var, {})
        ag_row = likert_ag.get(var, {})
        if not sa_row and not ag_row:
            continue
        try:
            overall_agree = round(
                float(sa_row.get("overall", 0) or 0) +
                float(ag_row.get("overall", 0) or 0), 1
            )
        except (ValueError, TypeError):
            overall_agree = 0
        agree_var = var + "_agree"
        for cid in cluster_ids:
            try:
                pct = round(
                    float(sa_row.get(f"c{cid}", 0) or 0) +
                    float(ag_row.get(f"c{cid}", 0) or 0), 1
                )
                result[cid][agree_var] = {
                    "pct": pct,
                    "question": meta["question"],
                    "domain": meta["domain"],
                    "diffPp": round(pct - overall_agree, 1),
                    "overall": round(overall_agree, 1),
                }
            except (ValueError, TypeError):
                pass

    # Phase 3: pew_churatd — weekly+ church attendance (needed by PartyCard)
    church_more = {}
    church_once = {}
    for r in rows:
        if r["variable"] != "pew_churatd":
            continue
        lbl = r.get("stat_label", "")
        if "More than once/week" in lbl:
            church_more = r
        elif "Once/week" in lbl and "More" not in lbl:
            church_once = r
    if church_more or church_once:
        try:
            overall_church = round(
                float(church_more.get("overall", 0) or 0) +
                float(church_once.get("overall", 0) or 0), 1
            )
        except (ValueError, TypeError):
            overall_church = 0
        for cid in cluster_ids:
            try:
                pct = round(
                    float(church_more.get(f"c{cid}", 0) or 0) +
                    float(church_once.get(f"c{cid}", 0) or 0), 1
                )
                result[cid]["pew_churatd"] = {
                    "pct": pct,
                    "question": "Weekly church attendance",
                    "domain": "Religion",
                    "diffPp": round(pct - overall_church, 1),
                    "overall": round(overall_church, 1),
                }
            except (ValueError, TypeError):
                pass

    # Phase 3.5: trust → % "Not very much" + % "None at all" (= % low trust)
    trust_meta_cv: dict = {}
    trust_not_much_cv: dict = {}
    trust_none_cv: dict = {}
    for r in rows:
        var = r["variable"]
        if r.get("type") == "trust":
            trust_meta_cv[var] = {"question": r.get("question", var), "domain": r.get("domain", "")}
        elif r.get("type") == "trust_dist":
            lbl = r.get("stat_label", "")
            if "Not very much" in lbl:
                trust_not_much_cv[var] = r
            elif "None at all" in lbl:
                trust_none_cv[var] = r

    TRUST_Q_CV = {
        "CC24_423": "Low trust in federal government (not very much or none at all)",
        "CC24_424": "Low trust in state government (not very much or none at all)",
    }
    for var, meta in trust_meta_cv.items():
        nm_row = trust_not_much_cv.get(var, {})
        na_row = trust_none_cv.get(var, {})
        if not nm_row and not na_row:
            continue
        try:
            overall = round(
                float(nm_row.get("overall", 0) or 0) +
                float(na_row.get("overall", 0) or 0), 1
            )
        except (ValueError, TypeError):
            overall = 0
        question = TRUST_Q_CV.get(var, meta["question"])
        for cid in cluster_ids:
            try:
                pct = round(
                    float(nm_row.get(f"c{cid}", 0) or 0) +
                    float(na_row.get(f"c{cid}", 0) or 0), 1
                )
                result[cid][var] = {
                    "pct": pct,
                    "question": question,
                    "domain": meta["domain"],
                    "diffPp": round(pct - overall, 1),
                    "overall": round(overall, 1),
                }
            except (ValueError, TypeError):
                pass

    # Phase 5: categorical_dist variables mapped to new demographic domains
    # (race/gender breakdowns are all enumerated below; no generic "Party base is X" leftover phase)
    CAT_INCLUSIONS = [
        # (variable, stat_label, new_domain, new_question, synth_key)
        # --- Household ---
        ('marstat',   '% Married',        'Household', 'Married',                   'marstat_married'),
        ('marstat',   '% Never married',  'Household', 'Never married',             'marstat_never'),
        ('ownhome',   '% Own',            'Household', 'Owns home',                 'ownhome_own'),
        ('ownhome',   '% Rent',           'Household', 'Rents home',                'ownhome_rent'),
        ('urbancity', '% City',           'Household', 'Lives in: city',            'urbancity_city'),
        ('urbancity', '% Suburb',         'Household', 'Lives in: suburb',          'urbancity_suburb'),
        ('urbancity', '% Town',           'Household', 'Lives in: town/small city', 'urbancity_town'),
        ('urbancity', '% Rural area',     'Household', 'Lives in: rural area',      'urbancity_rural'),
        # --- Race & Ethnicity ---
        ('race', '% White',                       'Race & Ethnicity', 'White',                           'race_white'),
        ('race', '% Black',                       'Race & Ethnicity', 'Black',                           'race_black'),
        ('race', '% Hispanic',                    'Race & Ethnicity', 'Hispanic',                        'race_hispanic'),
        ('race', '% Asian',                       'Race & Ethnicity', 'Asian',                           'race_asian'),
        ('race', '% Native American',             'Race & Ethnicity', 'Native American',                 'race_native_american'),
        ('race', '% Multiracial',                 'Race & Ethnicity', 'Multiracial',                     'race_multiracial'),
        ('immstat', '% Immigrant, naturalized',   'Race & Ethnicity', 'Immigrant (naturalized citizen)', 'immstat_nat'),
        ('immstat', '% Immigrant, not citizen',   'Race & Ethnicity', 'Immigrant (not yet a citizen)',   'immstat_nc'),
        ('immstat', '% US-born, parent immigrant','Race & Ethnicity', 'US-born, parent was immigrant',   'immstat_parent'),
        # --- Gender & Sexuality ---
        ('gender4',   '% Man',                   'Gender & Sexuality', 'Identifies as man',              'gender4_man'),
        ('gender4',   '% Woman',                 'Gender & Sexuality', 'Identifies as woman',            'gender4_woman'),
        ('gender4',   '% Non-binary',            'Gender & Sexuality', 'Non-binary or other gender',     'gender4_nonbinary'),
        ('sexuality', '% Heterosexual/straight', 'Gender & Sexuality', 'Heterosexual / straight',        'sexuality_het'),
        ('sexuality', '% Lesbian/gay woman',     'Gender & Sexuality', 'Lesbian',                        'sexuality_lesbian'),
        ('sexuality', '% Gay man',               'Gender & Sexuality', 'Gay man',                        'sexuality_gay'),
        ('sexuality', '% Bisexual',              'Gender & Sexuality', 'Bisexual',                       'sexuality_bisexual'),
        # --- Education ---
        ('educ', '% No HS',         'Education', 'Less than high school',        'educ_no_hs'),
        ('educ', '% HS grad',       'Education', 'High school graduate',         'educ_hs'),
        ('educ', '% Some college',  'Education', 'Some college (no degree)',     'educ_some_college'),
        ('educ', '% 2-year degree', 'Education', "Associate's degree (2-year)",  'educ_2yr'),
        ('educ', '% 4-year degree', 'Education', "Bachelor's degree (4-year)",   'educ_4yr'),
        ('educ', '% Post-grad',     'Education', 'Post-graduate degree',         'educ_postgrad'),
        # --- Economics: employment ---
        ('employ', '% Full-time',  'Economics', 'Employed full-time',   'employ_ft'),
        ('employ', '% Part-time',  'Economics', 'Employed part-time',   'employ_pt'),
        ('employ', '% Unemployed', 'Economics', 'Currently unemployed', 'employ_unemployed'),
        ('employ', '% Retired',    'Economics', 'Retired',              'employ_retired'),
        ('employ', '% Homemaker',  'Economics', 'Homemaker',            'employ_homemaker'),
        # --- Economics: union membership ---
        ('union',   '% Current member',   'Economics', 'Current union member',       'union_current'),
        ('union',   '% Former member',    'Economics', 'Former union member',        'union_former'),
        ('unionhh', '% Currently member', 'Economics', 'Household has union member', 'unionhh_current'),
        # --- Voting History: now built directly from the CES by pipeline/add_compare_items.py
        #     (2016/2020/2024 vote shares incl. third-party sums + Biden/Harris approval) ---
        # --- Other ---
        ('gunown', '% No one in HH',    'Other', 'No gun in household',   'gunown_none'),
        ('gunown', '% Personally owns', 'Other', 'Personally owns a gun', 'gunown_personal'),
    ]

    cat_lookup = {(r['variable'], r.get('stat_label', '')): r for r in rows}
    for var, stat_lbl, domain, question, synth_key in CAT_INCLUSIONS:
        r = cat_lookup.get((var, stat_lbl))
        if not r:
            continue
        try:
            overall = float(r.get('overall', 0) or 0)
        except (ValueError, TypeError):
            overall = 0
        for cid in cluster_ids:
            try:
                val = float(r.get(f'c{cid}', 0) or 0)
                result[cid][synth_key] = {
                    'pct': round(val, 1),
                    'question': question,
                    'domain': domain,
                    'diffPp': round(val - overall, 1),
                    'overall': round(overall, 1),
                }
            except (ValueError, TypeError):
                pass

    # Phase 5b: income tier groupings from faminc_new ordinal distribution
    INCOME_TIERS = [
        ('income_under50k', 'Family income under $50k',
         ['% <$10k', '% $10k\u201320k', '% $20k\u201330k', '% $30k\u201340k', '% $40k\u201350k']),
        ('income_50k_100k', 'Family income $50k\u2013$100k',
         ['% $50k\u201360k', '% $60k\u201370k', '% $70k\u201380k', '% $80k\u2013100k']),
        ('income_over100k', 'Family income over $100k',
         ['% $100k\u2013120k', '% $120k\u2013150k', '% $150k\u2013200k', '% $200k\u2013250k',
          '% $250k\u2013350k', '% $350k\u2013500k', '% $500k+']),
    ]
    inc_rows = {r['stat_label']: r for r in rows if r['variable'] == 'faminc_new'}
    for synth_key, question, labels in INCOME_TIERS:
        matching = [inc_rows[lbl] for lbl in labels if lbl in inc_rows]
        if not matching:
            continue
        try:
            overall = sum(float(r.get('overall', 0) or 0) for r in matching)
        except Exception:
            overall = 0
        for cid in cluster_ids:
            try:
                val = sum(float(r.get(f'c{cid}', 0) or 0) for r in matching)
                result[cid][synth_key] = {
                    'pct': round(val, 1),
                    'question': question,
                    'domain': 'Economics',
                    'diffPp': round(val - overall, 1),
                    'overall': round(overall, 1),
                }
            except Exception:
                pass

    # Phase 5c: remap existing binary variables to new domains
    BINARY_REMAP = {
        'gigwork':   ('Economics',       'Gig / freelance worker'),
        'investor':  ('Economics',       'Owns stocks or mutual funds'),
        'child18':   ('Household',       'Has children under 18'),
        'CC24_323f': ('Taxes & Economy', 'Forgive up to $20k of student loan debt per person'),
        'milstat_1': ('Other',           'Currently serving in the military'),
        'milstat_3': ('Other',           'Previously served in the military'),
    }
    for cid in cluster_ids:
        for var, (new_domain, new_question) in BINARY_REMAP.items():
            if var in result[cid]:
                result[cid][var]['domain'] = new_domain
                result[cid][var]['question'] = new_question

    # Phase 6: continuous variables with custom axis scale
    CONTINUOUS_VARS = [
        # (variable, stat_label, synth_key, domain, question, maxVal, unit)
        ('CC24_325', 'Median', 'CC24_325_median', 'Abortion',
         'Median abortion cutoff (weeks)', 40, 'wks'),
    ]
    for var, stat_lbl, synth_key, domain, question, max_val, unit in CONTINUOUS_VARS:
        r = cat_lookup.get((var, stat_lbl))
        if not r:
            continue
        try:
            overall = float(r.get('overall', 0) or 0)
        except (ValueError, TypeError):
            overall = 0
        for cid in cluster_ids:
            try:
                val = float(r.get(f'c{cid}', 0) or 0)
                result[cid][synth_key] = {
                    'pct': round(val, 1),
                    'question': question,
                    'domain': domain,
                    'diffPp': round(val - overall, 1),
                    'overall': round(overall, 1),
                    'maxVal': max_val,
                    'unit': unit,
                }
            except (ValueError, TypeError):
                pass

    return result

# "What sets a party apart" should surface political/social *views*, not demographics
# or voting behavior.
_NON_POLICY_DOMAINS = {"Voting History", "Demographics", "Employment & Labor"}


def compute_key_positions(rows, cid, n=4):
    """Return top-n data-driven policy positions that most differentiate this cluster."""
    binary_rows = [r for r in rows if r["type"] == "binary"
                   and r.get("domain") not in _NON_POLICY_DOMAINS]
    diffs = []
    for r in binary_rows:
        try:
            overall = float(r["overall"])
            val = float(r[f"c{cid}"]) if r.get(f"c{cid}") else overall
            diff = val - overall
            diffs.append((abs(diff), diff, r["question"], val))
        except (ValueError, KeyError):
            pass
    diffs.sort(reverse=True)
    out = []
    for _, diff, question, pct in diffs[:n]:
        out.append({
            "question": question,
            "pct": round(pct, 1),
            "direction": "supports" if diff > 0 else "opposes",
            "diffPp": round(diff, 1),
        })
    return out


def compute_key_positions_vs_neighbors(rows, cid, cluster_factors, n=4, min_diff=15):
    """Return top-n positions most distinguishing this cluster from its 2 nearest neighbors.
    Falls back to overall-diff approach if fewer than n positions pass the threshold."""
    me = cluster_factors.get(cid)
    if not me:
        return compute_key_positions(rows, cid, n)

    factor_keys = ["F1", "F2", "F3", "F4", "F5"]
    distances = []
    for other_cid, other in cluster_factors.items():
        if other_cid == cid:
            continue
        dist = sum((me[f] - other[f]) ** 2 for f in factor_keys) ** 0.5
        distances.append((dist, other_cid))
    distances.sort()
    neighbor_cids = [c2 for _, c2 in distances[:2]]

    binary_rows = [r for r in rows if r["type"] == "binary"
                   and r.get("domain") not in _NON_POLICY_DOMAINS]
    diffs = []
    for r in binary_rows:
        try:
            val = float(r[f"c{cid}"])
            neighbor_vals = [float(r[f"c{nc}"]) for nc in neighbor_cids if r.get(f"c{nc}")]
            if not neighbor_vals:
                continue
            avg_neighbor = sum(neighbor_vals) / len(neighbor_vals)
            diff = val - avg_neighbor
            if abs(diff) >= min_diff:
                diffs.append((abs(diff), diff, r["question"], val))
        except (ValueError, KeyError):
            pass
    diffs.sort(reverse=True)
    out = []
    for _, diff, question, pct in diffs[:n]:
        out.append({
            "question": question,
            "pct": round(pct, 1),
            "direction": "supports" if diff > 0 else "opposes",
            "diffPp": round(diff, 1),
        })
    # Fall back to overall-diff for any remaining slots
    if len(out) < n:
        seen = {p["question"] for p in out}
        fallback = compute_key_positions(rows, cid, n * 2)
        for pos in fallback:
            if pos["question"] not in seen:
                out.append(pos)
                seen.add(pos["question"])
            if len(out) >= n:
                break
    return out

def _compute_cluster_factor_centroids() -> dict:
    """Weighted-mean FS_F1..F5 centroid for every cluster 0-9.
    Returns {cid_str: {'F1':..,'F5':..}} — the same definition the coalition
    profiles used, computed directly so it doesn't depend on pruned inputs."""
    efa_path = Path(__file__).parent.parent.parent / "data" / "processed" / "efa_factor_scores.csv"
    typo_path = Path(__file__).parent.parent.parent / "data" / "processed" / "typology_cluster_assignments.csv"
    efa_rows = read_csv(str(efa_path))
    typo_rows = read_csv(str(typo_path))
    factor_map = {"F1": "FS_F1", "F2": "FS_F2", "F3": "FS_F3", "F4": "FS_F4", "F5": "FS_F5"}
    if efa_rows and "FS_F4" not in efa_rows[0]:
        factor_map["F4"] = "FS_F4_resid"
        factor_map["F5"] = "FS_F5_resid"

    def _val(row_idx, col):
        v = efa_rows[row_idx].get(col)
        if v is None or v == "":
            v = typo_rows[row_idx].get(col)
        return float(v or 0)

    num = {str(k): {fk: 0.0 for fk in factor_map} for k in range(10)}
    den = {str(k): 0.0 for k in range(10)}
    for i, tr in enumerate(typo_rows):
        cid = str(tr.get("cluster", "")).strip().replace(".0", "")
        if cid not in den:
            continue
        w = float(efa_rows[i].get("commonpostweight", 1) or 0)
        den[cid] += w
        for fk, col in factor_map.items():
            num[cid][fk] += _val(i, col) * w
    return {cid: {fk: round(num[cid][fk] / den[cid], 4) if den[cid] else 0.0
                  for fk in factor_map} for cid in num}


def build_cluster_profiles(include_c7=True, out_name="clusterProfiles.json",
                           house_summary_csv=None):
    rows = read_csv(OUTPUTS / "profiles" / "cluster_stats.csv")
    clusters = {str(i): {"id": str(i), "variables": {}}
                for i in range(10) if include_c7 or str(i) != "7"}

    all_vars = collect_cluster_variables(rows, include_c7=include_c7)
    for cid in clusters:
        clusters[cid]["variables"] = all_vars.get(cid, {})

    # Party code + name from the canonical cluster→party map; factor centroids
    # computed directly from EFA (independent of the pruned coalition diagnostic).
    centroids = _compute_cluster_factor_centroids()
    cluster_factors = {}
    for cid in clusters:
        code = CLUSTER_TO_PARTY.get(cid, cid)
        clusters[cid]["party"] = code
        clusters[cid]["partyName"] = PARTY_NAMES.get(code, code)
        fac = centroids.get(cid, {f"F{k}": 0.0 for k in range(1, 6)})
        for fk, v in fac.items():
            clusters[cid][fk] = v
        cluster_factors[cid] = dict(fac)

    # Seat counts: source from the pure-multi (party-line) house run — the same run
    # the House tab shows — so the party cards match the rest of the app (not the
    # coalition diagnostic, which uses a different apportionment).
    pm_house = OUTPUTS / "pure_multi" / "house" / "stv_seat_summary.csv"
    if pm_house.exists():
        for r in read_csv(pm_house):
            cid = str(int(r["party"]))
            if cid in clusters:
                clusters[cid]["seatsHouse"] = int(r["NATIONAL"])

    # Compute percentile ranks AND z-scores (SDs from population mean)
    efa_path = Path(__file__).parent.parent.parent / "data" / "processed" / "efa_factor_scores.csv"
    typo_path = Path(__file__).parent.parent.parent / "data" / "processed" / "typology_cluster_assignments.csv"
    if efa_path.exists() and typo_path.exists():
        efa_rows = read_csv(str(efa_path))
        typo_rows = read_csv(str(typo_path))
        factor_map = {"F1": "FS_F1", "F2": "FS_F2", "F3": "FS_F3", "F4": "FS_F4", "F5": "FS_F5"}
        if "FS_F4" not in (efa_rows[0] if efa_rows else {}):
            factor_map["F4"] = "FS_F4_resid"
            factor_map["F5"] = "FS_F5_resid"
        def get_factor_val(row_idx, factor_key):
            col = factor_map[factor_key]
            val = efa_rows[row_idx].get(col)
            if val is None or val == '':
                val = typo_rows[row_idx].get(col)
            return float(val or 0)

        N = len(efa_rows)
        for factor_key in ["F1", "F2", "F3", "F4", "F5"]:
            all_vals = [get_factor_val(i, factor_key) for i in range(N)]
            pop_mean = sum(all_vals) / N
            pop_sd = (sum((v - pop_mean) ** 2 for v in all_vals) / N) ** 0.5
            for cid in clusters:
                centroid = clusters[cid].get(factor_key, 0)
                below = sum(1 for v in all_vals if v < centroid)
                pctile = round(below / N * 100, 1)
                z_mean = round((centroid - pop_mean) / pop_sd, 2) if pop_sd > 0 else 0
                z_zero = round(centroid / pop_sd, 2) if pop_sd > 0 else 0
                clusters[cid][f"pctile_{factor_key}"] = pctile
                clusters[cid][f"z_{factor_key}"] = z_zero  # zero-based (factor model origin)

    # Add key positions vs nearest neighbors
    for cid in clusters:
        clusters[cid]["keyPositions"] = compute_key_positions_vs_neighbors(rows, cid, cluster_factors)

    write_json(list(clusters.values()), out_name)


def _extract_policy_vars(rows, get_val, max_vars=None):
    """Build a variables dict from cluster_stats or blend_stats rows.
    Handles binary, binary_agree, and likert5 (summed as SA+A).
    get_val(row) -> float | None; return None to skip a row.
    """
    result = {}

    # Phase 1: binary + binary_agree
    for r in rows:
        typ = r.get("type", "")
        lbl = r.get("stat_label", "")
        if not ((typ == "binary" and lbl == "% Supporting") or
                (typ == "binary_agree" and lbl == "% Agreeing")):
            continue
        val = get_val(r)
        if val is None:
            continue
        try:
            overall = float(r.get("overall") or 0)
            result[r["variable"]] = {
                "pct": round(val, 1),
                "question": r.get("question", r["variable"]),
                "domain": r.get("domain", ""),
                "diffPp": round(val - overall, 1),
                "overall": round(overall, 1),
            }
        except (ValueError, KeyError):
            pass

    # Phase 2: likert5 → % Strongly Agree + % Agree
    likert_meta: dict = {}
    likert_sa: dict = {}
    likert_ag: dict = {}
    for r in rows:
        var = r["variable"]
        if r.get("type") == "likert5":
            likert_meta[var] = {"question": r.get("question", var), "domain": r.get("domain", "")}
        elif r.get("type") == "likert5_dist":
            if r.get("stat_label") == "% Strongly Agree":
                likert_sa[var] = r
            elif r.get("stat_label") == "% Agree":
                likert_ag[var] = r

    for var, meta in likert_meta.items():
        sa_row = likert_sa.get(var, {})
        ag_row = likert_ag.get(var, {})
        sa_val = get_val(sa_row) if sa_row else None
        ag_val = get_val(ag_row) if ag_row else None
        if sa_val is None and ag_val is None:
            continue
        try:
            val = (sa_val or 0) + (ag_val or 0)
            overall = (float(sa_row.get("overall") or 0) if sa_row else 0) + \
                      (float(ag_row.get("overall") or 0) if ag_row else 0)
            result[var + "_agree"] = {
                "pct": round(val, 1),
                "question": meta["question"],
                "domain": meta["domain"],
                "diffPp": round(val - overall, 1),
            }
        except (ValueError, TypeError):
            pass

    # Phase 3: trust → % "Not very much" + % "None at all" (= % low trust)
    trust_meta: dict = {}
    trust_not_much: dict = {}
    trust_none_at_all: dict = {}
    for r in rows:
        var = r["variable"]
        if r.get("type") == "trust":
            trust_meta[var] = {"question": r.get("question", var), "domain": r.get("domain", "")}
        elif r.get("type") == "trust_dist":
            lbl = r.get("stat_label", "")
            if "Not very much" in lbl:
                trust_not_much[var] = r
            elif "None at all" in lbl:
                trust_none_at_all[var] = r

    TRUST_Q = {
        "CC24_423": "Low trust in federal government (not very much or none at all)",
        "CC24_424": "Low trust in state government (not very much or none at all)",
    }
    for var, meta in trust_meta.items():
        nm_row = trust_not_much.get(var, {})
        na_row = trust_none_at_all.get(var, {})
        nm_val = get_val(nm_row) if nm_row else None
        na_val = get_val(na_row) if na_row else None
        if nm_val is None and na_val is None:
            continue
        try:
            val = (nm_val or 0) + (na_val or 0)
            overall = (float(nm_row.get("overall") or 0) if nm_row else 0) +                       (float(na_row.get("overall") or 0) if na_row else 0)
            result[var] = {
                "pct": round(val, 1),
                "question": TRUST_Q.get(var, meta["question"]),
                "domain": meta["domain"],
                "diffPp": round(val - overall, 1),
            }
        except (ValueError, TypeError):
            pass

    items = sorted(result.items(), key=lambda x: abs(x[1]["diffPp"]), reverse=True)
    return dict(items[:max_vars] if max_vars else items)


# ---------- quizQuestions.json ----------
# Quiz variables: (variable, factor, row_selection_strategy)
# row_selection_strategy:
#   "binary"       → use type=="binary" row (% Supporting)
#   "likert_agree" → sum % Strongly Agree + % Agree rows
#   "trust_none"   → use % None at all row
#   "church_never" → use % Never row
QUIZ_VARS = [
    ("CC24_323b", "F1", "binary"),
    ("CC24_321d", "F1", "binary"),
    ("CC24_421_1", "F2", "likert_agree"),
    ("CC24_421_2", "F2", "likert_agree"),
    ("CC24_423",   "F3", "trust_none"),
    ("CC24_424",   "F3", "trust_none"),
    ("CC24_324b",  "F4", "binary"),
    ("pew_churatd","F4", "church_never"),
    ("CC24_323a",  "F5", "binary"),
    ("CC24_440a",  "F5", "likert_agree"),
]

QUIZ_QUESTION_OVERRIDES = {
    'CC24_423': 'I have very little trust in the federal government.',
    'CC24_424': 'I have very little trust in my state government.',
    'pew_churatd': 'I rarely or never attend religious services.',
}

CIDS = ["0","1","2","3","4","5","6","8","9"]

def build_quiz():
    rows = read_csv(OUTPUTS / "profiles" / "cluster_stats.csv")

    # Group rows by variable
    by_var = defaultdict(list)
    for r in rows:
        by_var[r["variable"]].append(r)

    questions = []
    for var, factor, strategy in QUIZ_VARS:
        var_rows = by_var.get(var, [])
        if not var_rows:
            print(f"  WARNING: quiz var {var} not found in cluster_stats.csv")
            continue

        cluster_pcts = {}
        question_text = QUIZ_QUESTION_OVERRIDES.get(var)
        domain = var_rows[0].get("domain", "")

        if strategy == "binary":
            row = next((r for r in var_rows if r.get("type") == "binary"), var_rows[0])
            if not question_text:
                question_text = row.get("question", var)
            for cid in CIDS:
                val = row.get(f"c{cid}", "")
                cluster_pcts[cid] = float(val) / 100 if val else 0.5

        elif strategy == "likert_agree":
            sa_row = next((r for r in var_rows if r.get("stat_label", "") == "% Strongly Agree"), None)
            a_row  = next((r for r in var_rows if r.get("stat_label", "") == "% Agree"), None)
            if not question_text:
                question_text = var_rows[0].get("question", var)
            for cid in CIDS:
                sa = float(sa_row.get(f"c{cid}", 0) or 0) if sa_row else 0
                a  = float(a_row.get(f"c{cid}", 0)  or 0) if a_row  else 0
                cluster_pcts[cid] = round((sa + a) / 100, 4)

        elif strategy == "trust_none":
            row = next((r for r in var_rows if "None at all" in r.get("stat_label", "")), var_rows[-1])
            if not question_text:
                question_text = row.get("question", var)
            for cid in CIDS:
                val = row.get(f"c{cid}", "")
                cluster_pcts[cid] = float(val) / 100 if val else 0.5

        elif strategy == "church_never":
            row = next((r for r in var_rows if r.get("stat_label", "") == "% Never"), var_rows[-1])
            if not question_text:
                question_text = row.get("question", var)
            for cid in CIDS:
                val = row.get(f"c{cid}", "")
                cluster_pcts[cid] = float(val) / 100 if val else 0.5

        questions.append({
            "variable": var,
            "factor": factor,
            "question": question_text or var,
            "domain": domain,
            "clusterSupport": cluster_pcts,
        })

    write_json(questions, "quizQuestions.json")


# ══════════════════════════════════════════════════════════════════════════════
# Factor Deviation (FD) builders
# ══════════════════════════════════════════════════════════════════════════════

# ---------- fdSenateCondorcet.json + fdSenateIRV.json ----------
_FD_PARTY_CLUSTER = {
    "CON": "0", "LBR": "1", "STY": "2", "NAT": "3", "LIB": "4",
    "POP": "5", "CUP": "6", "DSA": "8", "PRG": "9", "OAO": "7",
}

def build_fd_senate(src_dir=FD_DIR, cond_name="fdSenateCondorcet.json", irv_name="fdSenateIRV.json"):
    """FD senate compositions. CSV uses senator_party/axis/dir instead of label/cluster cols."""
    def _extract(rows):
        out = []
        for r in rows:
            party = r["senator_party"]
            out.append({
                "stateFips":   r["state_fips"].zfill(2),
                "stateAbbr":   r["state_abbr"],
                "senatorCode": r["senator_code"],
                "senatorParty": party,
                "senatorAxis": r["senator_axis"],
                "senatorDir":  r["senator_dir"],
                # SenateSeat-compatible fields for SenateMap / ParliamentChart
                "senatorLabel":     r["senator_code"],
                "senatorType":      "",
                "primaryCluster":   _FD_PARTY_CLUSTER.get(party, ""),
                "secondaryCluster": "",
            })
        return out

    cond_rows = read_csv(src_dir / "senate" / "senate_composition.csv")
    irv_rows  = read_csv(src_dir / "senate" / "senate_irv_composition.csv")
    write_json(_extract(cond_rows), cond_name)
    write_json(_extract(irv_rows),  irv_name)


# ---------- fdHouseSeats.json ----------
def build_fd_house_seats(src_dir=FD_DIR, out_name="fdHouseSeats.json"):
    """FD house seat summary. CSV uses candidate_code + party/axis/direction strings."""
    rows = read_csv(src_dir / "house" / "stv_seat_summary.csv")
    out = []
    for r in rows:
        out.append({
            "code":        r["candidate_code"],
            "party":       r["party"],
            "axis":        r["axis"],
            "direction":   r["direction"],
            "urban":       int(r["URBAN"]),
            "suburban":    int(r["SUBURBAN"]),
            "rural":       int(r["RURAL"]),
            "national":    int(r["NATIONAL"]),
            "pctNational": float(r["pct_national"]),
        })
    write_json(out, out_name)



def _build_trajectory_pcts(diag_rows):
    """Per-(candidate_code, stage) first-choice % from optional 'trajectories'
    diagnostic rows. The current primary diagnostics are transfer-format, so this
    is normally empty and callers fall back to the per-stage vote_pct in
    primary_results — which is the correct trajectory value."""
    out = {}
    for r in diag_rows:
        if r.get("diagnostic") != "trajectories":
            continue
        try:
            out[(r.get("candidate_code", ""), r.get("phase", ""))] = float(r.get("vote_pct") or 0)
        except (ValueError, TypeError):
            pass
    return out


# ---------- fdPrimary.json ----------
def build_fd_primary():
    """Stage-by-stage data for the 71-candidate FD primary run."""
    rows       = read_csv(FD_DIR / "primary_results_2028.csv")
    centroids  = {r["candidate_code"]: r for r in read_csv(FD_DIR / "candidate_factor_centroids.csv")}
    diag_rows  = read_csv(FD_DIR / "primary_diagnostics_2028.csv")
    traj_pcts  = _build_trajectory_pcts(diag_rows)

    stages_order = ["After_Retail", "After_Pod_A", "After_Pod_C", "After_Pod_BD"]
    stage_labels = {
        "After_Retail": "Retail + Bench States",
        "After_Pod_A":      "After Pod A (West)",
        "After_Pod_C":      "After Pod C (South)",
        "After_Pod_BD":     "After Pods B+D (Final)",
    }

    by_candidate  = defaultdict(dict)
    quota_by_stage = {}
    for row in rows:
        stage = row["winnowing_point"]
        code  = row["candidate_code"]
        by_candidate[code][stage] = {
            "voteTotal":      float(row["vote_total"]),
            "votePct":        traj_pcts.get((code, stage), float(row["vote_pct"])),
            "status":         row["status"],
            "quotaThreshold": float(row["quota_threshold"]),
        }
        quota_by_stage[stage] = float(row["quota_threshold"])

    candidates = []
    for code, stages in by_candidate.items():
        c = centroids.get(code, {})
        candidates.append({
            "code":      code,
            "name":      code,
            "party":     c.get("party", ""),
            "axis":      c.get("axis", "base"),
            "direction": c.get("direction", "base"),
            "F1": float(c.get("F1_security_order", 0)),
            "F2": float(c.get("F2_electoral_skepticism", 0)),
            "F3": float(c.get("F3_government_distrust", 0)),
            "F4": float(c.get("F4_religious_traditionalism", 0)),
            "F5": float(c.get("F5_populist_conservatism", 0)),
            "stages": {s: stages.get(s, {
                "voteTotal": 0, "votePct": 0,
                "status": "previously_eliminated",
                "quotaThreshold": quota_by_stage.get(s, 0),
            }) for s in stages_order},
        })

    write_json({
        "stagesOrder":  stages_order,
        "stageLabels":  stage_labels,
        "quotaByStage": quota_by_stage,
        "candidates":   candidates,
    }, "fdPrimary.json")


# ---------- fdPrimaryStateWinners.json ----------
def build_fd_primary_state_winners():
    """State-level IRV winners from the FD presidential general."""
    irv_path = FD_DIR / "irv" / "irv_presidential_states_2028.csv"
    pod_rows = read_csv(FD_DIR / "state_pod_assignments.csv")
    pod_by_fips = {r["state_fips"].zfill(2): r["pod"] for r in pod_rows}

    state_rows = read_csv(irv_path)
    out = {}
    for r in state_rows:
        fips       = r["state_fips"].zfill(2)
        winner     = r.get("irv_winner", r.get("winner_code", ""))
        runner_up  = r.get("irv_runner_up", r.get("runner_up_code", ""))
        r1_cols    = [k for k in r.keys() if k.startswith("r1_pct_") or k.startswith("fc_pct_")]
        shares = {}
        for col in r1_cols:
            code = col.replace("r1_pct_", "").replace("fc_pct_", "")
            val  = float(r.get(col) or 0)
            if val > 0:
                shares[code] = val
        total = sum(shares.values())
        if total > 0:
            shares = {k: round(v / total, 4) for k, v in shares.items()}
        n_resp = int(float(r.get("n_respondents", 0) or r.get("total_weight", 0) or 0))
        out[fips] = {
            "stateAbbr":    r["state_abbr"],
            "winnerCode":   winner,
            "runnerUpCode": runner_up,
            "pod":          pod_by_fips.get(fips, "D"),
            "nRespondents": n_resp,
            "shares":       shares,
        }
    write_json(out, "fdPrimaryStateWinners.json")


# ---------- pureMultiPrimaryStateWinners.json ----------
def build_pure_multi_primary_state_winners():
    """State-level IRV winners from the Raw Multi primary."""
    irv_path = PURE_MULTI_DIR / "irv" / "irv_presidential_states_2028.csv"
    pod_rows = read_csv(PURE_MULTI_DIR / "state_pod_assignments.csv")
    pod_by_fips = {r["state_fips"].zfill(2): r["pod"] for r in pod_rows}

    state_rows = read_csv(irv_path)
    out = {}
    for r in state_rows:
        fips       = r["state_fips"].zfill(2)
        winner     = r.get("irv_winner", r.get("winner_code", ""))
        runner_up  = r.get("irv_runner_up", r.get("runner_up_code", ""))
        r1_cols    = [k for k in r.keys() if k.startswith("r1_pct_") or k.startswith("fc_pct_")]
        shares = {}
        for col in r1_cols:
            code = col.replace("r1_pct_", "").replace("fc_pct_", "")
            val  = float(r.get(col) or 0)
            if val > 0:
                shares[code] = val
        total = sum(shares.values())
        if total > 0:
            shares = {k: round(v / total, 4) for k, v in shares.items()}
        n_resp = int(float(r.get("n_respondents", 0) or r.get("total_weight", 0) or 0))
        out[fips] = {
            "stateAbbr":    r["state_abbr"],
            "winnerCode":   winner,
            "runnerUpCode": runner_up,
            "pod":          pod_by_fips.get(fips, "D"),
            "nRespondents": n_resp,
            "shares":       shares,
        }
    write_json(out, "pureMultiPrimaryStateWinners.json")


# ---------- fdPrimarySankey.json ----------
def build_fd_primary_sankey():
    """Sankey/alluvial data for the 71-candidate FD primary.
    Stage 0 first-choice proportions come from state_candidate_profiles.csv.
    """
    diag_rows = read_csv(FD_DIR / "primary_diagnostics_2028.csv")
    profiles  = read_csv(FD_DIR / "state_candidate_profiles.csv")

    # Stage 0: aggregate first-choice totals across states
    fc_totals: dict = {}
    for row in profiles:
        n = float(row.get("total_weighted_respondents") or 0)
        for k, v in row.items():
            if k.startswith("first_choice_"):
                code = k.replace("first_choice_", "")
                fc_totals[code] = fc_totals.get(code, 0) + n * float(v or 0)
    total_fc = sum(fc_totals.values()) or 1
    fc_pct   = {code: round(v / total_fc * 100, 3) for code, v in fc_totals.items()}

    stage_order  = ["After_Retail", "After_Pod_A", "After_Pod_C", "After_Pod_BD"]
    stage_to_idx = {s: i + 1 for i, s in enumerate(stage_order)}

    traj_rows    = [r for r in diag_rows if r.get("diagnostic") == "trajectories"]
    active_at    = {i: [] for i in range(1, 5)}
    vote_pct_at  = {}
    for r in traj_rows:
        stage_idx = stage_to_idx.get(r.get("phase"))
        if stage_idx is None:
            continue
        code = r["candidate_code"]
        pct  = float(r.get("vote_pct") or 0)
        if r.get("status") in ("active", "surviving", "elected") and pct > 0:
            active_at[stage_idx].append(code)
            vote_pct_at[(code, stage_idx)] = pct

    # Collect ALL transfers (surplus + elimination), keyed by (stage_idx, eliminated_code)
    all_xfers = {i: {} for i in range(1, 5)}
    for r in diag_rows:
        if r.get("diagnostic") != "transfer_analysis":
            continue
        stage_idx = stage_to_idx.get(r.get("winnowing_point"))
        if stage_idx is None:
            continue
        e_code = r.get("eliminated_code", "")
        d_code = r.get("dest_code", "")
        pct    = float(r.get("pct_of_eliminated_total") or 0)
        xtype  = r.get("transfer_type", "elimination")
        if e_code not in all_xfers[stage_idx]:
            all_xfers[stage_idx][e_code] = []
        all_xfers[stage_idx][e_code].append((d_code, pct, xtype))

    nodes = []
    for code, pct in sorted(fc_pct.items(), key=lambda x: -x[1]):
        nodes.append({"id": f"{code}__0", "label": code, "stageIdx": 0, "pct": round(pct, 3)})
    for stage_idx in range(1, 5):
        for code in active_at[stage_idx]:
            pct = vote_pct_at.get((code, stage_idx), 0)
            nodes.append({"id": f"{code}__{stage_idx}", "label": code, "stageIdx": stage_idx, "pct": round(pct, 3)})

    # Pre-compute exhausted (ballot-exhaustion) totals so we can add Exhausted sink nodes
    def _exhausted_votes(code, src_pct, dst_active_set, stage_xfers):
        if code in dst_active_set:
            return 0.0
        xfer_sum = sum(p for _, p, _ in stage_xfers.get(code, []))
        return src_pct * max(0.0, 1.0 - xfer_sum / 100)

    exhausted_at = {i: 0.0 for i in range(1, 5)}
    for code, pct in fc_pct.items():
        exhausted_at[1] += _exhausted_votes(code, pct, set(active_at[1]), all_xfers[1])
    for dst_idx in range(2, 5):
        da = set(active_at[dst_idx])
        for code in active_at[dst_idx - 1]:
            exhausted_at[dst_idx] += _exhausted_votes(code, vote_pct_at.get((code, dst_idx - 1), 0), da, all_xfers[dst_idx])
    for stage_idx in range(1, 5):
        if exhausted_at[stage_idx] > 0.01:
            nodes.append({"id": f"exhausted__{stage_idx}", "label": "Exhausted", "stageIdx": stage_idx, "pct": round(exhausted_at[stage_idx], 3)})

    node_ids = {n["id"] for n in nodes}
    links = []

    def add_link(src_id, tgt_id, value, xfer_type="continuation"):
        if src_id in node_ids and tgt_id in node_ids and value > 0.01:
            links.append({"source": src_id, "target": tgt_id, "value": round(value, 3), "type": xfer_type})

    def draw_links_for_code(code, src_id, src_pct, dst_idx, dst_active, stage_xfers):
        """Draw continuation (retained) + surplus/elimination/exhausted outgoing links."""
        xfers = stage_xfers.get(code, [])
        in_dst = code in dst_active
        if in_dst:
            surplus_out = sum(p for _, p, t in xfers if t == "surplus") / 100
            retained = src_pct * (1.0 - surplus_out)
            add_link(src_id, f"{code}__{dst_idx}", retained, "continuation")
            for dest_code, xfer_pct, _ in xfers:
                add_link(src_id, f"{dest_code}__{dst_idx}", src_pct * xfer_pct / 100, "surplus")
        else:
            xfer_sum = sum(p for _, p, _ in xfers)
            for dest_code, xfer_pct, _ in xfers:
                add_link(src_id, f"{dest_code}__{dst_idx}", src_pct * xfer_pct / 100, "elimination")
            exhausted = src_pct * max(0.0, 1.0 - xfer_sum / 100)
            if exhausted > 0.01:
                add_link(src_id, f"exhausted__{dst_idx}", exhausted, "exhausted")

    # Stage 0 → 1
    retail_active = set(active_at[1])
    for code, pct in fc_pct.items():
        draw_links_for_code(code, f"{code}__0", pct, 1, retail_active, all_xfers[1])

    # Stages 1 → 2 → 3 → 4
    for dst_idx in range(2, 5):
        src_idx   = dst_idx - 1
        dst_active = set(active_at[dst_idx])
        for code in active_at[src_idx]:
            src_pct = vote_pct_at.get((code, src_idx), 0)
            draw_links_for_code(code, f"{code}__{src_idx}", src_pct, dst_idx, dst_active, all_xfers[dst_idx])

    n_survivors = [len(fc_pct), len(active_at[1]), len(active_at[2]), len(active_at[3]), len(active_at[4])]
    stage_labels = [
        f"Initial Slate ({n_survivors[0]})",
        f"After Retail ({n_survivors[1]})",
        f"After Pod A ({n_survivors[2]})",
        f"After Pod C ({n_survivors[3]})",
        f"Final ({n_survivors[4]})",
    ]
    write_json({"stageLabels": stage_labels, "nodes": nodes, "links": links}, "fdPrimarySankey.json")


# ---------- fdPresidentialElection.json ----------
def build_fd_presidential_election():
    """FD general election: national IRV rounds, Condorcet matchups, per-state winners."""
    irv_dir = FD_DIR / "irv"

    irv_rows      = read_csv(irv_dir / "irv_presidential_national_2028.csv")
    rounds_by_num = defaultdict(list)
    for r in irv_rows:
        rounds_by_num[int(r["round"])].append(r)

    irv_rounds = []
    irv_winner = None
    for rnum in sorted(rounds_by_num.keys()):
        candidates = []
        for r in rounds_by_num[rnum]:
            code      = r["candidate_code"]
            eliminated = r["eliminated"].strip().lower() == "true"
            winner     = r["winner"].strip().lower() == "true"
            if winner and not eliminated:
                irv_winner = code
            candidates.append({
                "code":      code,
                "name":      r["candidate_name"],
                "pct":       round(float(r["vote_pct"]), 2),
                "votes":     round(float(r["vote_total"]), 0),
                "eliminated": eliminated,
                "winner":    winner,
            })
        irv_rounds.append({"round": rnum, "candidates": candidates})

    condorcet_matchups = []
    condorcet_winner   = None
    cond_rows = read_csv(irv_dir / "condorcet_matchups_2028.csv")
    for r in cond_rows:
        a       = r["candidate_a"]
        b       = r["candidate_b"]
        votes_a = float(r["votes_a_beats_b"])
        votes_b = float(r["votes_b_beats_a"])
        total   = votes_a + votes_b
        condorcet_matchups.append({
            "candidateA": a,
            "candidateB": b,
            "aWinsPct":   round(votes_a / total * 100, 3) if total > 0 else 50.0,
            "margin":     round(float(r["margin_pct"]), 3),
            "winner":     r["winner"],
        })
        if not condorcet_winner and r.get("condorcet_winner"):
            condorcet_winner = r["condorcet_winner"]

    state_rows  = read_csv(irv_dir / "irv_presidential_states_2028.csv")
    pod_rows    = read_csv(FD_DIR / "state_pod_assignments.csv")
    pod_by_fips = {r["state_fips"].zfill(2): r["pod"] for r in pod_rows}

    irv_state_winners = {}
    for r in state_rows:
        fips    = r["state_fips"].zfill(2)
        r1_cols = [k for k in r.keys() if k.startswith("r1_pct_")]
        shares  = {}
        for col in r1_cols:
            code = col.replace("r1_pct_", "")
            val  = float(r.get(col) or 0)
            if val > 0:
                shares[code] = val
        total = sum(shares.values())
        irv_state_winners[fips] = {
            "stateAbbr":        r["state_abbr"],
            "winner":           r["winner_code"],
            "condorcetWinner":  r.get("condorcet_winner_code", r["winner_code"]),
            "pod":              pod_by_fips.get(fips, "D"),
            "nRespondents":     int(r["n_respondents"]),
            "shares":           {k: round(v / total, 4) for k, v in shares.items()} if total > 0 else {},
        }

    write_json({
        "irvRounds":        irv_rounds,
        "irvWinner":        irv_winner,
        "condorcetMatchups": condorcet_matchups,
        "condorcetWinner":  condorcet_winner,
        "irvStateWinners":  irv_state_winners,
    }, "fdPresidentialElection.json")


# ---------- rawMultiPresidentialElection.json ----------
def build_raw_multi_presidential_election(src_dir=PURE_MULTI_DIR,
                                          out_name="rawMultiPresidentialElection.json"):
    """Raw Multi general election: national IRV rounds, Condorcet matchups, per-state winners."""
    irv_dir = src_dir / "irv"

    irv_rows      = read_csv(irv_dir / "irv_presidential_national_2028.csv")
    rounds_by_num = defaultdict(list)
    for r in irv_rows:
        rounds_by_num[int(r["round"])].append(r)

    irv_rounds = []
    irv_winner = None
    for rnum in sorted(rounds_by_num.keys()):
        candidates = []
        for r in rounds_by_num[rnum]:
            code       = r.get("candidate_code", r.get("candidate", ""))
            status     = r.get("status", "surviving")
            eliminated = (r.get("eliminated", "").strip().lower() == "true") or (status == "eliminated_this_round")
            winner     = (r.get("winner", "").strip().lower() == "true") or (status == "winner")
            if winner and not eliminated:
                irv_winner = code
            candidates.append({
                "code":       code,
                "name":       r.get("candidate_name", r.get("party", code)),
                "pct":        round(float(r.get("vote_pct", r.get("pct", 0)) or 0), 2),
                "votes":      round(float(r.get("vote_total", r.get("votes", 0)) or 0), 0),
                "eliminated": eliminated,
                "winner":     winner,
            })
        irv_rounds.append({"round": rnum, "candidates": candidates})

    condorcet_matchups = []
    condorcet_winner   = None
    cond_rows = read_csv(irv_dir / "condorcet_matchups_2028.csv")
    for r in cond_rows:
        a       = r.get("candidate_a", r.get("a", ""))
        b       = r.get("candidate_b", r.get("b", ""))
        votes_a = float(r.get("votes_a_beats_b", r.get("votes_a", 0)) or 0)
        votes_b = float(r.get("votes_b_beats_a", r.get("votes_b", 0)) or 0)
        total   = votes_a + votes_b
        margin_pct = float(r.get("margin_pct", r.get("margin", 0)) or 0)
        if total > 0 and "margin_pct" not in r:
            margin_pct = (votes_a - votes_b) / total * 100
        condorcet_matchups.append({
            "candidateA": a,
            "candidateB": b,
            "aWinsPct":   round(votes_a / total * 100, 3) if total > 0 else 50.0,
            "margin":     round(margin_pct, 3),
            "winner":     a if votes_a >= votes_b else b,
        })
        if not condorcet_winner and r.get("condorcet_winner"):
            condorcet_winner = r["condorcet_winner"]

    state_rows  = read_csv(irv_dir / "irv_presidential_states_2028.csv")
    pod_path    = src_dir / "state_pod_assignments.csv"
    if not pod_path.exists():
        pod_path = PURE_MULTI_DIR / "state_pod_assignments.csv"   # pods are party-independent
    pod_rows    = read_csv(pod_path) if pod_path.exists() else []
    pod_by_fips = {r["state_fips"].zfill(2): r["pod"] for r in pod_rows}

    irv_state_winners = {}
    for r in state_rows:
        fips    = str(r["state_fips"]).zfill(2)
        fc_cols = [k for k in r.keys() if k.startswith("r1_pct_") or k.startswith("fc_pct_")]
        shares  = {}
        for col in fc_cols:
            code = col.replace("r1_pct_", "").replace("fc_pct_", "")
            val  = float(r.get(col) or 0)
            if val > 0:
                shares[code] = val
        total = sum(shares.values())
        winner = r.get("winner_code", r.get("irv_winner", ""))
        n_resp = int(float(r.get("n_respondents", 0) or r.get("total_weight", 0) or 0))
        irv_state_winners[fips] = {
            "stateAbbr":        r["state_abbr"],
            "winner":           winner,
            "condorcetWinner":  r.get("condorcet_winner_code", winner),
            "pod":              pod_by_fips.get(fips, "D"),
            "nRespondents":     n_resp,
            "shares":           {k: round(v / total, 4) for k, v in shares.items()} if total > 0 else {},
        }

    write_json({
        "irvRounds":         irv_rounds,
        "irvWinner":         irv_winner,
        "condorcetMatchups": condorcet_matchups,
        "condorcetWinner":   condorcet_winner,
        "irvStateWinners":   irv_state_winners,
    }, out_name)


# ---------- fdProfiles.json ----------
def build_fd_profiles():
    """Position profiles for all 71 FD candidates. Key positions = items where
    |candidate_value - overall| >= 3pp (binary/likert rows only).
    """
    stats_path   = FD_DIR / "profiles" / "factor_deviation_stats.csv"
    centroid_rows = read_csv(FD_DIR / "candidate_factor_centroids.csv")
    centroids     = {r["candidate_code"]: r for r in centroid_rows}
    cand_codes    = [r["candidate_code"] for r in centroid_rows]

    try:
        stat_rows = read_csv(stats_path)
    except FileNotFoundError:
        print(f"    [WARN] {stats_path} not found")
        write_json({}, "fdProfiles.json")
        return

    profiles = {}
    for code in cand_codes:
        c = centroids.get(code, {})
        key_positions = []
        for row in stat_rows:
            if row.get("type") not in ("binary", "likert"):
                continue
            try:
                val     = float(row.get(code, 0) or 0)
                overall = float(row.get("overall", 0) or 0)
                diff    = val - overall
            except (ValueError, TypeError):
                continue
            if abs(diff) >= 3.0:
                key_positions.append({
                    "variable": row["variable"],
                    "question": row.get("question", ""),
                    "domain":   row.get("domain", ""),
                    "value":    round(val, 1),
                    "overall":  round(overall, 1),
                    "diff":     round(diff, 1),
                })
        key_positions.sort(key=lambda x: -abs(x["diff"]))

        # Build variables dict (top 40 by |diffPp|) for CompareTab.
        # Use _extract_policy_vars() so we correctly handle binary, binary_agree,
        # likert5 (SA+A), and trust (not_very_much + none_at_all) types.
        def _get_val(r, _code=code):
            try:
                return float(r.get(_code) or 0)
            except (ValueError, TypeError):
                return None
        variables = _extract_policy_vars(stat_rows, _get_val, max_vars=40)

        profiles[code] = {
            "code":      code,
            "party":     c.get("party", ""),
            "axis":      c.get("axis", "base"),
            "direction": c.get("direction", "base"),
            "F1": float(c.get("F1_security_order", 0)),
            "F2": float(c.get("F2_electoral_skepticism", 0)),
            "F3": float(c.get("F3_government_distrust", 0)),
            "F4": float(c.get("F4_religious_traditionalism", 0)),
            "F5": float(c.get("F5_populist_conservatism", 0)),
            "keyPositions": key_positions[:8],
            "variables": variables,
        }

    write_json(profiles, "fdProfiles.json")


# ---------- pureMultiPrimary.json ----------
def build_pure_multi_primary(src_dir=PURE_MULTI_DIR, out_name="pureMultiPrimary.json"):
    """Stage-by-stage data for the 21-candidate pure/raw multi primary run."""
    rows      = read_csv(src_dir / "primary_results_2028.csv")
    diag_rows = read_csv(src_dir / "primary_diagnostics_2028.csv")
    traj_pcts = _build_trajectory_pcts(diag_rows)

    # Load party centroids keyed by candidate_name (= party code: CON, SD, etc.)
    centroids_rows = read_csv(OUTPUTS / "candidate_factor_centroids.csv")
    party_f = {
        r["candidate_name"]: {
            "F1": float(r.get("F1_security_order", 0) or 0),
            "F2": float(r.get("F2_electoral_skepticism", 0) or 0),
            "F3": float(r.get("F3_government_distrust", 0) or 0),
            "F4": float(r.get("F4_religious_traditionalism", 0) or 0),
            "F5": float(r.get("F5_populist_conservatism", 0) or 0),
        }
        for r in centroids_rows
    }

    # Auto-detect stages from the CSV
    all_stages = []
    for row in rows:
        s = row.get("winnowing_point", row.get("stage", ""))
        if s and s not in all_stages:
            all_stages.append(s)
    # Filter to non-initial stages
    stages_order = [s for s in all_stages if s != "Initial_Slate"]
    stage_labels = {s: s.replace("_", " ").replace("After ", "") for s in stages_order}

    by_candidate   = defaultdict(dict)
    quota_by_stage = {}
    party_of_code  = {}
    for row in rows:
        stage = row.get("winnowing_point", row.get("stage", ""))
        code  = row.get("candidate_code", row.get("candidate", ""))
        party_of_code[code] = row.get("party_code", row.get("party", code.rsplit("_", 1)[0]))
        vote_pct = float(row.get("vote_pct", row.get("vote_share", 0)) or 0)
        vote_total = float(row.get("vote_total", 0) or 0)
        quota = float(row.get("quota_threshold", 0) or 0)
        by_candidate[code][stage] = {
            "voteTotal":      vote_total,
            "votePct":        traj_pcts.get((code, stage), vote_pct),
            "status":         row.get("status", "surviving"),
            "quotaThreshold": quota,
        }
        if quota > 0:
            quota_by_stage[stage] = quota

    candidates = []
    for code, stages in by_candidate.items():
        party = party_of_code.get(code, code.rsplit("_", 1)[0])
        f = party_f.get(party, {"F1": 0.0, "F2": 0.0, "F3": 0.0, "F4": 0.0, "F5": 0.0})
        candidates.append({
            "code":      code,
            "name":      code,
            "party":     party,
            "axis":      "base",
            "direction": "base",
            "F1": f["F1"], "F2": f["F2"], "F3": f["F3"], "F4": f["F4"], "F5": f["F5"],
            "stages": {s: stages.get(s, {
                "voteTotal": 0, "votePct": 0,
                "status": "previously_eliminated",
                "quotaThreshold": quota_by_stage.get(s, 0),
            }) for s in stages_order},
        })

    write_json({
        "stagesOrder":  stages_order,
        "stageLabels":  stage_labels,
        "quotaByStage": quota_by_stage,
        "candidates":   candidates,
    }, out_name)


# ---------- pureMultiPrimarySankey.json ----------
def build_pure_multi_primary_sankey():
    """Sankey/alluvial data from primary_results + primary_diagnostics CSVs."""
    results_rows = read_csv(PURE_MULTI_DIR / "primary_results_2028.csv")
    diag_rows    = read_csv(PURE_MULTI_DIR / "primary_diagnostics_2028.csv")

    # Discover stages from results
    all_stages = []
    for r in results_rows:
        s = r.get("winnowing_point", r.get("stage", ""))
        if s and s not in all_stages:
            all_stages.append(s)
    initial_stage = all_stages[0] if all_stages else "Initial_Slate"
    elim_stages   = [s for s in all_stages if s != initial_stage]
    n_stages      = len(elim_stages)

    # Build per-stage candidate data {stage_idx: {code: {pct, status, vote_total}}}
    stage_to_idx = {s: i for i, s in enumerate([initial_stage] + elim_stages)}
    candidates_at = {i: {} for i in range(len(all_stages))}

    for r in results_rows:
        stage   = r.get("winnowing_point", r.get("stage", ""))
        code    = r.get("candidate_code", r.get("candidate", ""))
        pct     = float(r.get("vote_pct", r.get("vote_share", 0)) or 0)
        vtotal  = float(r.get("vote_total", 0) or 0)
        status  = r.get("status", "surviving")
        idx     = stage_to_idx.get(stage)
        if idx is not None:
            candidates_at[idx][code] = {"pct": pct, "status": status, "vote_total": vtotal}

    # Build transfer flows per stage {stage_idx: {from_code: [(to_code, votes, type)]}}
    transfers_at = {i: defaultdict(list) for i in range(1, len(all_stages))}
    for r in diag_rows:
        stage  = r.get("winnowing_point", "")
        idx    = stage_to_idx.get(stage)
        if idx is None or idx == 0:
            continue
        e_code = r.get("eliminated_code", "")
        d_code = r.get("dest_code", "")
        votes  = float(r.get("transferred_votes", 0) or 0)
        xtype  = r.get("transfer_type", "elimination")
        if e_code and d_code:
            transfers_at[idx][e_code].append((d_code, votes, xtype))

    # Compute total votes for percentage normalization
    total_votes = sum(c["pct"] for c in candidates_at[0].values()) or 100.0

    # Build nodes
    nodes = []
    for code, info in sorted(candidates_at[0].items(), key=lambda x: -x[1]["pct"]):
        if info["pct"] > 0:
            nodes.append({"id": f"{code}__0", "label": code, "stageIdx": 0, "pct": round(info["pct"], 3)})

    for stage_idx in range(1, len(all_stages)):
        for code, info in candidates_at[stage_idx].items():
            if info["status"] in ("surviving", "elected", "active"):
                pct = max(info["pct"], 0.1)  # minimum visibility for zero-vote survivors
                nodes.append({"id": f"{code}__{stage_idx}", "label": code, "stageIdx": stage_idx, "pct": round(pct, 3)})

    # Add exhausted nodes where needed
    for stage_idx in range(1, len(all_stages)):
        prev_codes = set(candidates_at[stage_idx - 1].keys())
        curr_surv  = {c for c, info in candidates_at[stage_idx].items()
                      if info["status"] in ("surviving", "elected", "active")}
        eliminated = prev_codes - curr_surv
        if eliminated:
            # Check if there's unaccounted vote flow
            nodes.append({"id": f"exhausted__{stage_idx}", "label": "Exhausted", "stageIdx": stage_idx, "pct": 0})

    node_ids = {n["id"] for n in nodes}
    links    = []

    def add_link(src, tgt, val, xtype="continuation"):
        if src in node_ids and tgt in node_ids and val > 0.005:
            links.append({"source": src, "target": tgt, "value": round(val, 3), "type": xtype})

    def _resolve_dest(code, stage_xfers, surviving, visited=None):
        """Follow elimination chains to find surviving final destinations.
        Returns {surviving_code: fraction_of_input} (fractions sum to <= 1.0)."""
        if visited is None:
            visited = set()
        if code in surviving:
            return {code: 1.0}
        if code in visited:
            return {}
        visited.add(code)
        flows = stage_xfers.get(code, [])
        total = sum(v for _, v, _ in flows)
        if total <= 0:
            return {}  # exhausted
        result = {}
        for next_dest, vol, _ in flows:
            sub = _resolve_dest(next_dest, stage_xfers, surviving, set(visited))
            for k, v in sub.items():
                result[k] = result.get(k, 0) + (vol / total) * v
        return result

    # Build links between stages
    for stage_idx in range(1, len(all_stages)):
        prev = candidates_at[stage_idx - 1]
        curr_surv = {c for c, info in candidates_at[stage_idx].items()
                     if info["status"] in ("surviving", "elected", "active")}
        xfers = transfers_at[stage_idx]

        for code, info in prev.items():
            if info["pct"] <= 0:
                continue  # skip zero-vote candidates (e.g. _2/_3 at Initial_Slate)
            src_id  = f"{code}__{stage_idx - 1}"
            src_pct = info["pct"]

            if code in curr_surv:
                # Survivor: continuation link + surplus transfers out
                surplus_flows = [(d, v, t) for d, v, t in xfers.get(code, []) if t == "surplus"]
                if not surplus_flows:
                    add_link(src_id, f"{code}__{stage_idx}", src_pct, "continuation")
                else:
                    surplus_total = sum(v for _, v, _ in surplus_flows)
                    source_votes = info.get("vote_total", 0) or surplus_total
                    effective = max(surplus_total, source_votes)
                    surplus_frac = surplus_total / effective

                    add_link(src_id, f"{code}__{stage_idx}",
                             src_pct * (1.0 - surplus_frac), "continuation")
                    leaked = 0.0
                    for dest, vol, _ in surplus_flows:
                        link_val = src_pct * vol / effective
                        if dest in curr_surv:
                            add_link(src_id, f"{dest}__{stage_idx}", link_val, "surplus")
                        else:
                            # Trace through elimination chain to surviving destinations
                            resolved = _resolve_dest(dest, xfers, curr_surv)
                            for final, rfrac in resolved.items():
                                add_link(src_id, f"{final}__{stage_idx}",
                                         link_val * rfrac, "surplus")
                            leaked += link_val * (1.0 - sum(resolved.values()))
                    if leaked > 0.01:
                        add_link(src_id, f"exhausted__{stage_idx}", leaked, "exhausted")
            else:
                # Eliminated: distribute to destinations, resolving chains
                flows = xfers.get(code, [])
                total_out = sum(v for _, v, _ in flows) or 1.0
                accounted = 0.0
                for dest, vol, xtype in flows:
                    frac = vol / total_out * src_pct
                    if dest in curr_surv:
                        add_link(src_id, f"{dest}__{stage_idx}", frac, "elimination")
                        accounted += frac
                    else:
                        resolved = _resolve_dest(dest, xfers, curr_surv)
                        for final, rfrac in resolved.items():
                            add_link(src_id, f"{final}__{stage_idx}",
                                     frac * rfrac, "elimination")
                            accounted += frac * rfrac
                remaining = src_pct - accounted
                if remaining > 0.01:
                    add_link(src_id, f"exhausted__{stage_idx}", remaining, "exhausted")

    # Stage labels
    counts = [
        sum(1 for info in candidates_at[i].values()
            if info["status"] in ("surviving", "elected", "active"))
        for i in range(len(all_stages))
    ]
    stage_label_names = ["Initial Slate"] + [s.replace("_", " ").replace("After ", "") for s in elim_stages]
    stage_labels = [f"{name} ({counts[i]})" for i, name in enumerate(stage_label_names)]

    write_json({"stageLabels": stage_labels, "nodes": nodes, "links": links}, "pureMultiPrimarySankey.json")


# ---------- pureMultiPrimaryBuckets.json ----------
def build_pure_multi_primary_buckets(src_dir=PURE_MULTI_DIR, out_name="pureMultiPrimaryBuckets.json"):
    """Per-stage bucket composition: for each winner, where did their quota come from?"""
    results_rows = read_csv(src_dir / "primary_results_2028.csv")
    diag_rows    = read_csv(src_dir / "primary_diagnostics_2028.csv")

    # Parse results by stage
    by_stage = defaultdict(dict)   # stage → {code: {pct, vote_total, status, party}}
    all_stages = []
    for r in results_rows:
        stage = r.get("winnowing_point", "")
        code  = r.get("candidate_code", "")
        if stage and stage not in all_stages:
            all_stages.append(stage)
        by_stage[stage][code] = {
            "pct":      float(r.get("vote_pct", 0) or 0),
            "vtotal":   float(r.get("vote_total", 0) or 0),
            "fc_total": float(r.get("first_choice_total", 0) or 0),
            "status":   r.get("status", ""),
            "party":    r.get("party_code", code.rsplit("_", 1)[0]),
            "quota":    float(r.get("quota_threshold", 0) or 0),
        }

    initial_stage = all_stages[0]  # Initial_Slate
    elim_stages   = [s for s in all_stages if s != initial_stage]
    pool = sum(v["vtotal"] for v in by_stage[initial_stage].values())

    # Parse diagnostics: incoming transfers per (stage, dest)
    xfer_incoming = defaultdict(lambda: defaultdict(float))   # (stage, dest) → {src_party: votes}
    xfer_outgoing = defaultdict(lambda: defaultdict(float))   # (stage, src) → {dest_code: votes}
    for d in diag_rows:
        stage = d.get("winnowing_point", "")
        src   = d.get("eliminated_code", "")
        dest  = d.get("dest_code", "")
        votes = float(d.get("transferred_votes", 0) or 0)
        src_party = src.rsplit("_", 1)[0]
        xfer_incoming[(stage, dest)][src_party] += votes
        xfer_outgoing[(stage, src)][dest] += votes

    stages_out = []
    prev_stage = initial_stage

    for stage in elim_stages:
        stage_data = by_stage[stage]
        prev_data  = by_stage[prev_stage]
        quota_raw  = next((v["quota"] for v in stage_data.values() if v["quota"] > 0), 0)
        quota_pct  = quota_raw / pool * 100 if pool > 0 else 0

        winners   = []
        eliminated = []

        for code, info in sorted(stage_data.items(), key=lambda x: -x[1]["pct"]):
            party   = info["party"]
            # Use first-choice tally at THIS stage (not carried from previous)
            fc_raw = float(info.get("fc_total", 0) or 0)
            entering_pct = fc_raw / pool * 100 if pool > 0 else 0
            inc = xfer_incoming.get((stage, code), {})

            # Build source breakdown as list sorted by size
            all_sources = []
            for src_party, v in sorted(inc.items(), key=lambda x: -x[1]):
                pct = v / pool * 100
                if pct > 0.05:
                    all_sources.append({"party": src_party, "pct": round(pct, 2)})

            inc_total_pct = sum(s["pct"] for s in all_sources)
            total_pct = entering_pct + inc_total_pct
            overflow  = total_pct - quota_pct

            if info["status"] in ("surviving", "elected") and info["pct"] > 0:
                # Cap composition at quota: if entering >= quota, the survivor
                # filled quota entirely from own first-choice votes (no transfers
                # needed). If entering < quota, show only enough transfers to
                # fill the gap.
                capped_entering = min(entering_pct, quota_pct)
                gap = max(quota_pct - entering_pct, 0)
                capped_sources = []
                remaining_gap = gap
                for src in all_sources:
                    if remaining_gap <= 0:
                        break
                    used = min(src["pct"], remaining_gap)
                    capped_sources.append({"party": src["party"], "pct": round(used, 2)})
                    remaining_gap -= used

                winners.append({
                    "code":      code,
                    "party":     party,
                    "entering":  round(capped_entering, 2),
                    "sources":   capped_sources,
                    "total":     round(total_pct, 2),
                    "retained":  round(info["pct"], 2),
                    "overflow":  round(max(overflow, 0), 2),
                })
            elif info["status"] == "eliminated_this_round":
                # Where did this candidate's votes go?
                out = xfer_outgoing.get((stage, code), {})
                out_total = sum(out.values()) or 1.0
                dests = []
                for dest_code, v in sorted(out.items(), key=lambda x: -x[1]):
                    pct = v / out_total * 100
                    if pct > 1:
                        dests.append({"code": dest_code, "pct": round(pct, 1)})
                eliminated.append({
                    "code":     code,
                    "party":    party,
                    "entering": round(entering_pct, 2),
                    "sources":  all_sources,
                    "total":    round(total_pct, 2),
                    "dests":    dests,
                })

        label = stage.replace("_", " ").replace("After ", "")
        n_entering = sum(1 for v in prev_data.values()
                         if v["status"] in ("surviving", "elected"))
        stages_out.append({
            "name":       stage,
            "label":      label,
            "quota":      round(quota_pct, 2),
            "nEntering":  n_entering,
            "nWinners":   len(winners),
            "winners":    winners,
            "eliminated": eliminated,
        })
        prev_stage = stage

    write_json({"pool": round(pool, 2), "stages": stages_out}, out_name)


# ---------- fdPrimaryBuckets.json ----------
def build_fd_primary_buckets():
    """Per-stage bucket composition for Factor Dev primary."""
    results_rows = read_csv(FD_DIR / "primary_results_2028.csv")
    diag_rows    = read_csv(FD_DIR / "primary_diagnostics_2028.csv")

    # Parse results by stage
    by_stage = defaultdict(dict)
    all_stages = []
    for r in results_rows:
        stage = r.get("winnowing_point", "")
        code  = r.get("candidate_code", "")
        if stage and stage not in all_stages:
            all_stages.append(stage)
        by_stage[stage][code] = {
            "pct":      float(r.get("vote_pct", 0) or 0),
            "vtotal":   float(r.get("vote_total", 0) or 0),
            "fc_total": float(r.get("first_choice_total", 0) or 0),
            "status":   r.get("status", ""),
            "party":    r.get("party_code", code.split("_")[0]),
            "quota":    float(r.get("quota_threshold", 0) or 0),
            "pool":     float(r.get("accumulated_pool_size", 0) or 0),
        }

    elim_stages = [s for s in all_stages if by_stage[s]]

    # Parse transfer_analysis diagnostics
    xfer_incoming = defaultdict(lambda: defaultdict(float))
    xfer_outgoing = defaultdict(lambda: defaultdict(float))
    for d in diag_rows:
        if d.get("diagnostic") != "transfer_analysis":
            continue
        stage = d.get("winnowing_point", "")
        src   = d.get("eliminated_code", "")
        dest  = d.get("dest_code", "")
        votes = float(d.get("transferred_votes", 0) or 0)
        src_party = src.split("_")[0]
        xfer_incoming[(stage, dest)][src_party] += votes
        xfer_outgoing[(stage, src)][dest] += votes

    stages_out = []
    for stage in elim_stages:
        stage_data = by_stage[stage]
        # Use stage-specific pool (accumulated_pool_size grows as pods vote)
        pool = next((v["pool"] for v in stage_data.values() if v["pool"] > 0), 1.0)
        quota_raw  = next((v["quota"] for v in stage_data.values() if v["quota"] > 0), 0)
        quota_pct  = quota_raw / pool * 100 if pool > 0 else 0

        winners    = []
        eliminated = []

        for code, info in sorted(stage_data.items(), key=lambda x: -x[1]["pct"]):
            party = info["party"]
            # Use first-choice tally if available, else fall back to vtotal
            fc_raw = info.get("fc_total", 0) or info.get("vtotal", 0)
            entering_pct = fc_raw / pool * 100 if pool > 0 else 0
            inc = xfer_incoming.get((stage, code), {})

            all_sources = []
            for src_party, v in sorted(inc.items(), key=lambda x: -x[1]):
                pct = v / pool * 100
                if pct > 0.05:
                    all_sources.append({"party": src_party, "pct": round(pct, 2)})

            inc_total_pct = sum(s["pct"] for s in all_sources)
            total_pct = entering_pct + inc_total_pct

            if info["status"] in ("surviving", "elected", "active") and info["pct"] > 0:
                capped_entering = min(entering_pct, quota_pct)
                gap = max(quota_pct - entering_pct, 0)
                capped_sources = []
                remaining_gap = gap
                for src in all_sources:
                    if remaining_gap <= 0:
                        break
                    used = min(src["pct"], remaining_gap)
                    capped_sources.append({"party": src["party"], "pct": round(used, 2)})
                    remaining_gap -= used

                winners.append({
                    "code":      code,
                    "party":     party,
                    "entering":  round(capped_entering, 2),
                    "sources":   capped_sources,
                    "total":     round(total_pct, 2),
                    "retained":  round(quota_pct, 2),
                    "overflow":  round(max(total_pct - quota_pct, 0), 2),
                })
            elif info["status"] == "eliminated_this_round":
                out = xfer_outgoing.get((stage, code), {})
                out_total = sum(out.values()) or 1.0
                dests = []
                for dest_code, v in sorted(out.items(), key=lambda x: -x[1]):
                    pct = v / out_total * 100
                    if pct > 1:
                        dests.append({"code": dest_code, "pct": round(pct, 1)})
                eliminated.append({
                    "code":     code,
                    "party":    party,
                    "entering": round(entering_pct, 2),
                    "sources":  all_sources,
                    "total":    round(total_pct, 2),
                    "dests":    dests,
                })

        label = stage.replace("_", " ").replace("After ", "")
        n_entering = len(stage_data)
        stages_out.append({
            "name":       stage,
            "label":      label,
            "quota":      round(quota_pct, 2),
            "nEntering":  n_entering,
            "nWinners":   len(winners),
            "winners":    winners,
            "eliminated": eliminated,
        })

    write_json({"pool": round(pool, 2), "stages": stages_out}, "fdPrimaryBuckets.json")


# ---------- pureMultiSenate*.json ----------
# ---------- senateBuckets.json ----------
def build_senate_buckets():
    """Per-state and national-average bucket compositions for senate finalists."""
    bucket_rows  = read_csv(PURE_MULTI_DIR / "senate" / "senate_stv_buckets.csv")
    cond_rows    = read_csv(PURE_MULTI_DIR / "senate" / "senate_composition.csv")
    irv_rows     = read_csv(PURE_MULTI_DIR / "senate" / "senate_irv_composition.csv")

    # Winner lookups: {fips: winner_code}
    cond_winner = {r["state_fips"].zfill(2): r["senator_code"] for r in cond_rows}
    irv_winner  = {r["state_fips"].zfill(2): r["senator_code"]  for r in irv_rows}

    PARTIES = ["CON", "CUP", "DSA", "LIB", "NAT", "PRG", "POP", "LBR", "STY", "OAO"]

    # Build per-state finalist bucket data
    states = {}  # fips → {finalists: [...], condWinner, irvWinner}
    for r in bucket_rows:
        fips  = r["state_fips"].zfill(2)
        code  = r["finalist_code"]
        party = r["finalist_party"]
        fc    = float(r.get("first_choice_pct", 0) or 0)
        sources = []
        for p in PARTIES:
            val = float(r.get(f"inc_{p}", 0) or 0)
            if val > 0.05:
                sources.append({"party": p, "pct": round(val, 2)})
        sources.sort(key=lambda s: -s["pct"])

        if fips not in states:
            abbr = r.get("state_abbr", fips)
            states[fips] = {
                "fips": fips, "abbr": abbr,
                "condWinner": cond_winner.get(fips, ""),
                "irvWinner":  irv_winner.get(fips, ""),
                "finalists": [],
            }
        states[fips]["finalists"].append({
            "code": code, "party": party,
            "firstChoice": round(fc, 2),
            "sources": sources,
            "total": round(fc + sum(s["pct"] for s in sources), 2),
        })

    # Sort finalists by total descending within each state
    for st in states.values():
        st["finalists"].sort(key=lambda f: -f["total"])

    # Build national averages: for each winning party, average the bucket composition
    # across all states where that party's candidate wins (Condorcet)
    from collections import defaultdict as _dd
    party_buckets = _dd(lambda: {"fc_sum": 0.0, "inc_sums": _dd(float), "count": 0})
    for fips, st in states.items():
        winner_code = st["condWinner"]
        if not winner_code:
            continue
        winner_party = winner_code.split("_")[0]
        # Find the winner's finalist entry
        winner_fin = next((f for f in st["finalists"] if f["code"] == winner_code), None)
        if not winner_fin:
            continue
        pb = party_buckets[winner_party]
        pb["count"] += 1
        pb["fc_sum"] += winner_fin["firstChoice"]
        for s in winner_fin["sources"]:
            pb["inc_sums"][s["party"]] += s["pct"]

    averages = []
    for party in PARTIES:
        pb = party_buckets.get(party)
        if not pb or pb["count"] == 0:
            continue
        n = pb["count"]
        avg_sources = []
        for p, total in sorted(pb["inc_sums"].items(), key=lambda x: -x[1]):
            avg = total / n
            if avg > 0.1:
                avg_sources.append({"party": p, "pct": round(avg, 2)})
        avg_fc = round(pb["fc_sum"] / n, 2)
        averages.append({
            "party": party,
            "seats": n,
            "avgFirstChoice": avg_fc,
            "avgSources": avg_sources,
            "avgTotal": round(avg_fc + sum(s["pct"] for s in avg_sources), 2),
        })
    averages.sort(key=lambda a: -a["seats"])

    write_json({
        "states": states,
        "averages": averages,
    }, "senateBuckets.json")


# ---------- senateCondorcet.json ----------
def build_senate_condorcet():
    """National-average Condorcet matrix + per-state matchups for senate finalists."""
    cond_rows = read_csv(PURE_MULTI_DIR / "senate" / "senate_condorcet_results.csv")

    PARTIES = ["PRG", "LIB", "DSA", "LBR", "OAO", "STY", "CUP", "CON", "POP", "NAT"]

    # Per-state matchups (for drill-down)
    states = {}
    for r in cond_rows:
        fips = r["state_fips"].zfill(2)
        if fips not in states:
            states[fips] = {
                "abbr": r["state_abbr"],
                "winner": r.get("rp_winner_overall", ""),
                "matchups": [],
            }
        a_votes = float(r.get("votes_a_beats_b", 0))
        b_votes = float(r.get("votes_b_beats_a", 0))
        total = a_votes + b_votes
        states[fips]["matchups"].append({
            "candidateA": r["candidate_a"],
            "candidateB": r["candidate_b"],
            "aWinsPct": round(a_votes / total, 4) if total else 0.5,
            "margin": round(float(r.get("margin_pct", 0)), 2),
            "winner": r["candidate_a"] if a_votes > b_votes else r["candidate_b"],
        })

    # National aggregation at party level
    pair_data = {}  # (partyA, partyB) → {a_wins, total, margin_sum}
    for r in cond_rows:
        pa = r["candidate_a"].rsplit("_", 1)[0]
        pb = r["candidate_b"].rsplit("_", 1)[0]
        if pa == pb:
            continue
        a_votes = float(r.get("votes_a_beats_b", 0))
        b_votes = float(r.get("votes_b_beats_a", 0))
        margin  = float(r.get("margin_pct", 0))
        # Normalize so party pair key is always alphabetically ordered
        if pa > pb:
            pa, pb = pb, pa
            a_votes, b_votes = b_votes, a_votes
            margin = -margin
        key = f"{pa}|{pb}"
        if key not in pair_data:
            pair_data[key] = {"a_wins": 0, "total": 0, "margin_sum": 0.0}
        pd_entry = pair_data[key]
        pd_entry["total"] += 1
        pd_entry["margin_sum"] += margin
        if a_votes > b_votes:
            pd_entry["a_wins"] += 1

    # Build matrix as {rowParty: {colParty: {winRate, avgMargin, n}}}
    matrix = {}
    for key, pd_entry in pair_data.items():
        pa, pb = key.split("|")
        n = pd_entry["total"]
        if n == 0:
            continue
        a_wr = pd_entry["a_wins"] / n
        avg_m = pd_entry["margin_sum"] / n
        if pa not in matrix:
            matrix[pa] = {}
        if pb not in matrix:
            matrix[pb] = {}
        matrix[pa][pb] = {"winRate": round(a_wr, 3), "avgMargin": round(avg_m, 1), "n": n}
        matrix[pb][pa] = {"winRate": round(1 - a_wr, 3), "avgMargin": round(-avg_m, 1), "n": n}

    # Determine overall Condorcet winner (party with highest avg win rate)
    party_wr = {}
    for p in PARTIES:
        if p not in matrix:
            continue
        rates = [v["winRate"] for v in matrix[p].values()]
        party_wr[p] = sum(rates) / len(rates) if rates else 0
    overall_winner = max(party_wr, key=party_wr.get) if party_wr else ""

    write_json({
        "parties": PARTIES,
        "matrix": matrix,
        "overallWinner": overall_winner,
        "states": states,
    }, "senateCondorcet.json")


def build_pure_multi_senate(src_dir=PURE_MULTI_DIR,
                            cond_name="pureMultiSenateCondorcet.json",
                            irv_name="pureMultiSenateIRV.json"):
    """Pure multi senate compositions. senator_code is 'PARTY_N'; party from senator_party col."""
    def _extract(rows):
        out = []
        for r in rows:
            party = r.get("senator_party", r["senator_code"].rsplit("_", 1)[0])
            out.append({
                "stateFips":        r["state_fips"].zfill(2),
                "stateAbbr":        r["state_abbr"],
                "senatorCode":      r["senator_code"],
                "senatorParty":     party,
                "senatorAxis":      "base",
                "senatorDir":       "base",
                # SenateSeat-compatible fields for SenateMap / ParliamentChart
                "senatorLabel":     r["senator_code"],
                "senatorType":      "pure",
                "primaryCluster":   _FD_PARTY_CLUSTER.get(party, ""),
                "secondaryCluster": "",
            })
        return out

    cond_rows = read_csv(src_dir / "senate" / "senate_composition.csv")
    irv_rows  = read_csv(src_dir / "senate" / "senate_irv_composition.csv")
    write_json(_extract(cond_rows), cond_name)
    write_json(_extract(irv_rows),  irv_name)


def build_fptp_disproportionality():
    rows = read_csv(RESULTS / "post_recs" / "Current Congressional Inequality.csv")
    out = []
    for r in rows:
        if not r.get("State"):
            continue
        try:
            out.append({
                "state": r["State"],
                "totalSeats": int(r["Total Seats"]),
                "gopFptpSeats": int(r["GOP FPTP Seats"]),
                "demFptpSeats": int(r["DEM FPTP Seats"]),
                "gopPrSeats": int(r["GOP PR Seats"]),
                "demPrSeats": int(r["Dem PR Seast"]),
                "gopVotePct": float(r["GOP Vote %"].rstrip("%")),
                "demVotePct": float(r["DEM Vote %"].rstrip("%")),
                "fptpSeatDiff": float(r["FPTP Seat Diff (pos = GOP, neg = DEM)"].rstrip("%")),
            })
        except (ValueError, KeyError):
            pass
    write_json(out, "fptpDisproportionality.json")


def build_district_stv_results(src_csv=None, out_name="districtStvResults.json"):
    """Group district-level STV results by state FIPS for county tier map."""
    if src_csv is None:
        src_csv = OUTPUTS / "pure_multi" / "house" / "stv_results_by_district.csv"
    path = src_csv
    if not path.exists():
        print(f"  Skipping {out_name} (not found: {path})")
        write_json({}, out_name)
        return
    rows = read_csv(path)
    by_state: dict = {}
    for r in rows:
        state_fips = str(r["state_fips"]).zfill(2)
        elected = [r[f"elected_{i}"] for i in range(20) if r.get(f"elected_{i}")]
        entry = {
            "districtId":  r["district_id"],
            "densityTier": r["density_tier"],
            "seatCount":   int(r["seat_count"]),
            "elected":     elected,
            "nRespondents": int(r["n_respondents"]),
        }
        by_state.setdefault(state_fips, []).append(entry)
    write_json(by_state, out_name)


def build_fd_district_stv_results(src_dir=FD_DIR, out_name="fdDistrictStvResults.json"):
    """Group FD district-level STV results by state FIPS."""
    path = src_dir / "house" / "stv_results_by_district.csv"
    if not path.exists():
        write_json({}, out_name)
        return
    rows = read_csv(path)
    by_state: dict = {}
    for r in rows:
        state_fips = str(r["state_fips"]).zfill(2)
        elected = [r[f"elected_{i}"] for i in range(9) if r.get(f"elected_{i}")]
        # Map variant codes to base party for coloring
        elected_parties = [e.split("_")[0] if "_" in e and not e.split("_")[1].isdigit() else e.split("_")[0] for e in elected]
        entry = {
            "districtId":  r["district_id"],
            "densityTier": r["density_tier"],
            "seatCount":   int(r["seat_count"]),
            "elected":     elected_parties,
            "electedFull": elected,
            "nRespondents": int(r["n_respondents"]),
        }
        by_state.setdefault(state_fips, []).append(entry)
    write_json(by_state, out_name)


def build_district_county_map():
    """Map each district_id to its list of county FIPS5 strings."""
    path = Path(__file__).parent.parent.parent / "data" / "processed" / "county_to_district.csv"
    if not path.exists():
        print(f"  Skipping districtCountyMap (not found: {path})")
        write_json({}, "districtCountyMap.json")
        return
    rows = read_csv(path)
    result: dict = {}
    for r in rows:
        did   = r["district_id"]
        fips5 = str(r["county_fips5"]).zfill(5)
        result.setdefault(did, []).append(fips5)
    write_json(result, "districtCountyMap.json")


def build_county_tiers():
    """Map county FIPS → density tier using CDC NCHS Urban-Rural Classification (2013).

    Source: https://www.cdc.gov/nchs/data/data_acces_files/NCHSURCodes2013.xlsx
    Save as data/raw/NCHSURCodes2013.xlsx

    NCHS codes:
      1 = Large Central Metro (≥1M CBSA, principal city county) → URBAN
      2 = Large Fringe Metro (≥1M CBSA, outlying)              → URBAN
      3 = Medium Metro (250k–999k CBSA)                        → SUBURBAN
      4 = Small Metro (<250k CBSA)                             → SUBURBAN
      5 = Micropolitan                                         → RURAL
      6 = Noncore Rural                                        → RURAL
    """
    import openpyxl
    raw_path = Path(__file__).parent.parent.parent / "data" / "raw" / "NCHSURCodes2013.xlsx"
    if not raw_path.exists():
        print(f"  Skipping countyTiers (not found: {raw_path})")
        write_json({}, "countyTiers.json")
        return

    NCHS_TIER = {1:'URBAN', 2:'URBAN', 3:'SUBURBAN', 4:'SUBURBAN', 5:'RURAL', 6:'RURAL'}

    wb = openpyxl.load_workbook(raw_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    tiers = {}
    for row in rows[1:]:  # skip header
        fips = row[0]
        if fips is None:
            continue
        fips5 = str(int(fips)).zfill(5)
        state = fips5[:2]
        code = row[6]
        if state in ('60', '66', '69', '72', '78'):
            tiers[fips5] = 'RURAL'  # US territories not in NCHS
        else:
            tiers[fips5] = NCHS_TIER.get(code, 'RURAL')
    write_json(tiers, "countyTiers.json")


def build_rcv_results():
    """Assemble processed RCV race JSONs into a single grouped file.

    If data/outputs/rcv/ contains processed race JSONs (from process_rcv.py),
    merges them into the output. Otherwise preserves the existing rcvResults.json
    (which contains hardcoded summary data) unchanged.
    """
    rcv_dir = Path(__file__).parent.parent.parent / "data" / "outputs" / "rcv"
    if not rcv_dir.exists() or not list(rcv_dir.glob("*.json")):
        return  # Preserve existing hardcoded summary data

    result: dict = {"AK": [], "ME": []}
    for fpath in sorted(rcv_dir.glob("*.json")):
        try:
            with open(fpath) as f:
                race = json.load(f)
            state = race.get("state", "")
            if state in result:
                result[state].append(race)
        except Exception:
            pass
    for state in result:
        result[state].sort(key=lambda r: (-r.get("year", 0), r.get("office", "")))
    write_json(result, "rcvResults.json")


# ══════════════════════════════════════════════════════════════════════════════
# Triple Wyoming data builders
# ══════════════════════════════════════════════════════════════════════════════

def build_house_seats_triple(src_dir=PURE_MULTI_TRIPLE_DIR, out_name="houseSeatsTriple.json"):
    path = src_dir / "house" / "stv_seat_summary.csv"
    if not path.exists():
        print(f"  Skipping {out_name} (not found: {path})")
        write_json([], out_name)
        return
    rows = read_csv(path)
    pop_shares = _national_pop_shares_10()
    out = []
    for r in rows:
        cluster = int(r["party"])
        out.append({
            "party": cluster,
            "partyName": r["party_name"],
            "urban": int(r["URBAN"]),
            "suburban": int(r["SUBURBAN"]),
            "rural": int(r["RURAL"]),
            "national": int(r["NATIONAL"]),
            "pctNational": float(r["pct_national"]),
            "pctPopulation": pop_shares.get(cluster, 0.0),
        })
    write_json(out, out_name)


def build_fd_house_seats_triple(src_dir=FD_TRIPLE_DIR, out_name="fdHouseSeatsTriple.json"):
    path = src_dir / "house" / "stv_seat_summary.csv"
    if not path.exists():
        print(f"  Skipping {out_name} (not found: {path})")
        write_json([], out_name)
        return
    rows = read_csv(path)
    out = []
    for r in rows:
        out.append({
            "code":        r["candidate_code"],
            "party":       r["party"],
            "axis":        r["axis"],
            "direction":   r["direction"],
            "urban":       int(r["URBAN"]),
            "suburban":    int(r["SUBURBAN"]),
            "rural":       int(r["RURAL"]),
            "national":    int(r["NATIONAL"]),
            "pctNational": float(r["pct_national"]),
        })
    write_json(out, out_name)


def build_house_state_map_triple():
    path = PURE_MULTI_TRIPLE_DIR / "house" / "stv_results_by_district.csv"
    if not path.exists():
        print(f"  Skipping houseStateMapTriple (not found: {path})")
        write_json({}, "houseStateMapTriple.json")
        return
    rows = read_csv(path)
    pod_path = OUTPUTS / "state_pod_assignments.csv"
    if not pod_path.exists():
        pod_path = PURE_MULTI_DIR / "state_pod_assignments.csv"
    pod_rows = read_csv(pod_path) if pod_path.exists() else []
    abbr_by_fips = {r["state_fips"].zfill(2): r["state_abbr"] for r in pod_rows}

    state_seats = defaultdict(lambda: defaultdict(int))
    for row in rows:
        fips = row["state_fips"].zfill(2)
        for i in range(12):  # up to 10-seat districts
            v = row.get(f"elected_{i}", "").strip()
            if v:
                state_seats[fips][v] += 1

    state_pop_shares = _compute_state_pop_shares()

    out = {}
    for fips, counts in state_seats.items():
        total = sum(counts.values())
        plurality = max(counts, key=counts.get)
        out[fips] = {
            "stateAbbr": abbr_by_fips.get(fips, fips),
            "pluralityParty": plurality,
            "totalSeats": total,
            "seats": dict(counts),
            "popShares": state_pop_shares.get(fips, {}),
        }
    write_json(out, "houseStateMapTriple.json")


def build_district_stv_results_triple(src_dir=PURE_MULTI_TRIPLE_DIR, out_name="districtStvResultsTriple.json"):
    path = src_dir / "house" / "stv_results_by_district.csv"
    if not path.exists():
        print(f"  Skipping {out_name} (not found: {path})")
        write_json({}, out_name)
        return
    rows = read_csv(path)
    by_state: dict = {}
    for r in rows:
        state_fips = str(r["state_fips"]).zfill(2)
        elected = [r[f"elected_{i}"] for i in range(12) if r.get(f"elected_{i}")]
        entry = {
            "districtId":  r["district_id"],
            "densityTier": r["density_tier"],
            "seatCount":   int(r["seat_count"]),
            "elected":     elected,
            "nRespondents": int(r["n_respondents"]),
        }
        by_state.setdefault(state_fips, []).append(entry)
    write_json(by_state, out_name)


def build_fd_district_stv_results_triple(src_dir=FD_TRIPLE_DIR, out_name="fdDistrictStvResultsTriple.json"):
    path = src_dir / "house" / "stv_results_by_district.csv"
    if not path.exists():
        print(f"  Skipping {out_name} (not found: {path})")
        write_json({}, out_name)
        return
    rows = read_csv(path)
    by_state: dict = {}
    for r in rows:
        state_fips = str(r["state_fips"]).zfill(2)
        elected = [r[f"elected_{i}"] for i in range(12) if r.get(f"elected_{i}")]
        elected_parties = [e.split("_")[0] if "_" in e and not e.split("_")[1].isdigit() else e.split("_")[0] for e in elected]
        entry = {
            "districtId":  r["district_id"],
            "densityTier": r["density_tier"],
            "seatCount":   int(r["seat_count"]),
            "elected":     elected_parties,
            "electedFull": elected,
            "nRespondents": int(r["n_respondents"]),
        }
        by_state.setdefault(state_fips, []).append(entry)
    write_json(by_state, out_name)


def build_district_county_map_triple():
    path = Path(__file__).parent.parent.parent / "data" / "processed" / "county_to_district_triple.csv"
    if not path.exists():
        print(f"  Skipping districtCountyMapTriple (not found: {path})")
        write_json({}, "districtCountyMapTriple.json")
        return
    rows = read_csv(path)
    result: dict = {}
    for r in rows:
        did   = r["district_id"]
        fips5 = str(r["county_fips5"]).zfill(5)
        result.setdefault(did, []).append(fips5)
    write_json(result, "districtCountyMapTriple.json")


def _national_pop_shares_10() -> dict:
    """Soft-weighted national population share per cluster across all 10 clusters
    (sums to ~100). Used for WFP houseSeats pctPopulation."""
    efa_path = Path(__file__).parent.parent.parent / "data" / "processed" / "efa_factor_scores.csv"
    typo_path = Path(__file__).parent.parent.parent / "data" / "processed" / "typology_cluster_assignments.csv"
    efa_rows = read_csv(str(efa_path))
    typo_rows = read_csv(str(typo_path))
    total_w = sum(float(r.get("commonpostweight", 1) or 0) for r in efa_rows)
    shares = {}
    for k in range(10):
        col = f"prob_cluster_{k}"
        num = sum(float(typo_rows[i].get(col, 0) or 0) * float(efa_rows[i].get("commonpostweight", 1) or 0)
                  for i in range(len(typo_rows)))
        shares[k] = round(num / total_w * 100, 2) if total_w else 0.0
    return shares


def _cluster_by_var_support() -> dict:
    """{variable: cluster_stats_row} for binary '% Supporting' policy items."""
    cluster_rows = read_csv(OUTPUTS / "profiles" / "cluster_stats.csv")
    return {r["variable"]: r for r in cluster_rows
            if r.get("stat_label") == "% Supporting" and r.get("type") == "binary"}


def build_senate_vote_model_wfp(src, out_name="senateVoteModelWFP.json"):
    """senateVoteModelWFP.json — clone of senateVoteModel.json with only the
    Raw-Multi senate + president columns recomputed from the C7 run."""
    with open(DATA_OUT / "senateVoteModel.json", encoding="utf-8") as f:
        base = json.load(f)
    cbv = _cluster_by_var_support()

    rm_irv_winner = None
    for r in read_csv(src / "irv" / "irv_presidential_national_2028.csv"):
        if r.get("winner", "").strip() == "True":
            rm_irv_winner = r["candidate_code"]
    rm_cond_winner = None
    _cm = list(read_csv(src / "irv" / "condorcet_matchups_2028.csv"))
    if _cm:
        rm_cond_winner = _cm[0].get("condorcet_winner") or None

    def _party(c):
        return c.rsplit("_", 1)[0] if c else ""

    cond_seats, irv_seats = {}, {}
    for r in read_csv(src / "senate" / "senate_composition.csv"):
        p = r["senator_code"].rsplit("_", 1)[0]; cond_seats[p] = cond_seats.get(p, 0) + 1
    for r in read_csv(src / "senate" / "senate_irv_composition.csv"):
        p = r["senator_code"].rsplit("_", 1)[0]; irv_seats[p] = irv_seats.get(p, 0) + 1
    cond_res = _lf_prob_pass(cond_seats, cbv)
    irv_res  = _lf_prob_pass(irv_seats,  cbv)
    irv_party, cond_party = _party(rm_irv_winner), _party(rm_cond_winner)

    for row in base:
        var  = row["variable"]
        crow = cbv.get(var)
        if crow:
            irv_sup  = _lf_senator_support(irv_party,  crow)
            cond_sup = _lf_senator_support(cond_party, crow)
            row["presRawMultiIRVSigns"]  = "SIGN" if irv_sup  > 50 else "VETO"
            row["presRawMultiIRVPct"]    = round(irv_sup, 2)
            row["presRawMultiCondSigns"] = "SIGN" if cond_sup > 50 else "VETO"
            row["presRawMultiCondPct"]   = round(cond_sup, 2)
        row["condRawMultiProbPass"] = cond_res.get(var, {}).get("prob_pass", 0.0)
        row["condRawMultiVerdict"]  = cond_res.get(var, {}).get("verdict", "N/A")
        row["irvRawMultiProbPass"]  = irv_res.get(var, {}).get("prob_pass", 0.0)
        row["irvRawMultiVerdict"]   = irv_res.get(var, {}).get("verdict", "N/A")

    # ── Crossover (FD) senate columns from the native FD run (OAO-inclusive) ──
    fd_stats = read_csv(FD_DIR / "profiles" / "factor_deviation_stats.csv")
    fd_by_var = {r["variable"]: r for r in fd_stats
                 if r.get("stat_label") == "% Supporting" and r.get("type") == "binary"
                 and r.get("variable", "").startswith("CC24_")}
    cond_fd, irv_fd = {}, {}
    for r in read_csv(FD_DIR / "senate" / "senate_composition.csv"):
        cond_fd[r["senator_code"]] = cond_fd.get(r["senator_code"], 0) + 1
    for r in read_csv(FD_DIR / "senate" / "senate_irv_composition.csv"):
        irv_fd[r["senator_code"]] = irv_fd.get(r["senator_code"], 0) + 1
    fd_cond, fd_irv = _fd_prob_pass(cond_fd, fd_by_var), _fd_prob_pass(irv_fd, fd_by_var)
    fd_irv_w = None
    for r in read_csv(FD_DIR / "irv" / "irv_presidential_national_2028.csv"):
        if r.get("winner", "").strip() == "True":
            fd_irv_w = r["candidate_code"]
    _fcm = list(read_csv(FD_DIR / "irv" / "condorcet_matchups_2028.csv"))
    fd_cond_w = _fcm[0].get("condorcet_winner") if _fcm else None
    for row in base:
        var = row["variable"]; frow = fd_by_var.get(var)
        row["condFDProbPass"] = fd_cond.get(var, {}).get("prob_pass", 0.0)
        row["condFDVerdict"]  = fd_cond.get(var, {}).get("verdict", "N/A")
        row["irvFDProbPass"]  = fd_irv.get(var, {}).get("prob_pass", 0.0)
        row["irvFDVerdict"]   = fd_irv.get(var, {}).get("verdict", "N/A")
        if frow:
            isup = float(frow.get(fd_irv_w) or 0) if fd_irv_w else 0.0
            csup = float(frow.get(fd_cond_w) or 0) if fd_cond_w else 0.0
            row["presFDIRVSigns"]  = "SIGN" if isup > 50 else "VETO"; row["presFDIRVPct"]  = round(isup, 2)
            row["presFDCondSigns"] = "SIGN" if csup > 50 else "VETO"; row["presFDCondPct"] = round(csup, 2)
            row["presFDSigns"] = row["presFDIRVSigns"]; row["presFDPct"] = row["presFDIRVPct"]
    write_json(base, out_name)


def build_house_vote_model_wfp(src, out_name="houseVoteModelWFP.json"):
    """houseVoteModelWFP.json — clone of houseVoteModel.json with only the
    Raw-Multi house column recomputed from the C7 run (WFP seats included)."""
    with open(DATA_OUT / "houseVoteModel.json", encoding="utf-8") as f:
        base = json.load(f)
    cbv = _cluster_by_var_support()
    _cluster_to_party = {v: k for k, v in _PURE_CLUSTER.items()}

    rm_seats, total = {}, 0
    for r in read_csv(src / "house" / "stv_seat_summary.csv"):
        cl   = int(r["party"])
        code = _cluster_to_party.get(cl, str(cl))
        rm_seats[code] = rm_seats.get(code, 0) + int(r["NATIONAL"])
        total += int(r["NATIONAL"])
    maj = total // 2 + 1
    res = _lf_prob_pass(rm_seats, cbv, majority=maj)

    def _house_res(csv_path, by_code):
        seats, tot = {}, 0
        if not Path(csv_path).exists():
            return {}
        for r in read_csv(csv_path):
            code = r["party"] if by_code else _cluster_to_party.get(int(r["party"]), str(r["party"]))
            seats[code] = seats.get(code, 0) + int(r["NATIONAL"])
            tot += int(r["NATIONAL"])
        return _lf_prob_pass(seats, cbv, majority=tot // 2 + 1) if tot else {}

    fd_res  = _house_res(FD_DIR / "house" / "stv_seat_summary.csv", by_code=True)
    rmt_res = _house_res(PURE_MULTI_TRIPLE_DIR / "house" / "stv_seat_summary.csv", by_code=False)
    fdt_res = _house_res(FD_TRIPLE_DIR / "house" / "stv_seat_summary.csv", by_code=True)
    for row in base:
        var = row["variable"]
        row["houseRawMultiProbPass"] = res.get(var, {}).get("prob_pass", 0.0)
        row["houseRawMultiVerdict"]  = res.get(var, {}).get("verdict", "N/A")
        row["houseFDProbPass"] = fd_res.get(var, {}).get("prob_pass", 0.0)
        row["houseFDVerdict"]  = fd_res.get(var, {}).get("verdict", "N/A")
        if rmt_res:
            row["houseRawMultiTripleProbPass"] = rmt_res.get(var, {}).get("prob_pass", 0.0)
            row["houseRawMultiTripleVerdict"]  = rmt_res.get(var, {}).get("verdict", "N/A")
        if fdt_res:
            row["houseFDTripleProbPass"] = fdt_res.get(var, {}).get("prob_pass", 0.0)
            row["houseFDTripleVerdict"]  = fdt_res.get(var, {}).get("verdict", "N/A")
    write_json(base, out_name)


def build_nosty_scenario():
    """Parallel 'no-Solidarity' scenario: cluster 2 is dissolved and its voters' ballots
    flow to the remaining 9 parties. Reads the NO_STY=1 pipeline outputs (pure_multi_nosty/)
    and emits *NoSTY.json for the Presidency / Senate / House toggles."""
    d = PURE_MULTI_NOSTY_DIR
    build_raw_multi_presidential_election(src_dir=d, out_name="rawMultiPresidentialElectionNoSTY.json")
    build_pure_multi_senate(src_dir=d, cond_name="pureMultiSenateCondorcetNoSTY.json", irv_name="pureMultiSenateIRVNoSTY.json")
    build_house_seats(src_csv=d / "house" / "stv_seat_summary.csv", out_name="houseSeatsNoSTY.json")
    build_house_state_map(src_dir=d, out_name="houseStateMapNoSTY.json")
    build_district_stv_results(src_csv=d / "house" / "stv_results_by_district.csv", out_name="districtStvResultsNoSTY.json")
    # Legislation vote models — Raw-Multi chamber pass + president sign/veto recomputed from the no-STY run.
    build_senate_vote_model_wfp(d, out_name="senateVoteModelNoSTY.json")
    build_house_vote_model_wfp(d, out_name="houseVoteModelNoSTY.json")


def _build_turnout_variant(d, suffix):
    """Emit the *<suffix>.json family from a turnout-weighted pipeline tree `d`."""
    build_raw_multi_presidential_election(src_dir=d, out_name=f"rawMultiPresidentialElection{suffix}.json")
    build_pure_multi_senate(src_dir=d, cond_name=f"pureMultiSenateCondorcet{suffix}.json", irv_name=f"pureMultiSenateIRV{suffix}.json")
    build_house_seats(src_csv=d / "house" / "stv_seat_summary.csv", out_name=f"houseSeats{suffix}.json")
    build_house_state_map(src_dir=d, out_name=f"houseStateMap{suffix}.json")
    build_district_stv_results(src_csv=d / "house" / "stv_results_by_district.csv", out_name=f"districtStvResults{suffix}.json")
    build_senate_vote_model_wfp(d, out_name=f"senateVoteModel{suffix}.json")
    build_house_vote_model_wfp(d, out_name=f"houseVoteModel{suffix}.json")


def build_pure_multi_primary_state_shares(src_dir, out_name):
    """Per-stage, per-state first-choice shares (drives the primary national-share chart).
    Renames the pipeline's legacy SD_ candidate code to LBR_ to match the viz."""
    with open(src_dir / "primary_state_stage_shares.json", encoding="utf-8") as f:
        data = json.load(f)
    for st in data.values():
        for stg in st.get("stages", {}).values():
            shares = stg.get("shares", {})
            for k in list(shares.keys()):
                if k.startswith("SD_"):
                    shares["LBR_" + k[3:]] = shares.pop(k)
    write_json(data, out_name)


# Compression sweep stops (share of the inter-force turnout gap closed).
TURNOUT_STOPS = (10, 20, 30)


def build_turnout_scenario():
    """'Current participation' floor (λ=0): ballots weighted by validated 2024 turnout
    (TURNOUT_WEIGHT=1 pipeline). Emits *Turnout.json for president/senate/house/vote-models,
    plus the turnout-responsive primary (finalists, buckets, per-stage state shares)."""
    _build_turnout_variant(PURE_MULTI_TURNOUT_DIR, "Turnout")
    _build_turnout_variant(PURE_MULTI_NOSTY_TURNOUT_DIR, "NoStyTurnout")  # dormant (not wired)
    build_pure_multi_primary(PURE_MULTI_TURNOUT_DIR, "pureMultiPrimaryTurnout.json")
    build_pure_multi_primary_buckets(PURE_MULTI_TURNOUT_DIR, "pureMultiPrimaryBucketsTurnout.json")
    build_pure_multi_primary_state_shares(PURE_MULTI_TURNOUT_DIR, "pureMultiPrimaryStageSharesTurnout.json")


def build_party_population():
    """Each force's share of the adult population vs the as-cast 2024 electorate.
    Uses SOFT (GMM-posterior) weighting so popShare matches the app's canonical
    population share (_national_pop_shares_10 / House 'Population vs Seat Share');
    voteShare re-weights the same posteriors by validated 2024 vote."""
    import pandas as pd
    proc = Path(__file__).parent.parent.parent / "data" / "processed"
    efa  = pd.read_csv(proc / "efa_factor_scores.csv")
    typo = pd.read_csv(proc / "typology_cluster_assignments.csv")
    tp   = pd.read_csv(proc / "turnout_propensity.csv")
    w    = efa["commonpostweight"].values.astype(float)
    cluster = tp["cluster"].values                     # hard argmax cluster
    t_cluster = tp["turnout_cluster"].values           # per-cluster validated turnout (what the sim uses)
    CODES = ["CON", "LBR", "STY", "NAT", "LIB", "POP", "CUP", "OAO", "DSA", "PRG"]
    Wtot = w.sum()
    # popShare: soft posterior (canonical, matches _national_pop_shares_10 / House page).
    # turnout: hard per-cluster rate the compression sim weights by. voteShare: their product.
    pops, turns = [], []
    for k in range(len(CODES)):
        pk = typo[f"prob_cluster_{k}"].values.astype(float)
        pops.append(float((pk * w).sum()) / Wtot * 100)
        km = cluster == k
        turns.append(float(t_cluster[km][0]) if km.any() else 0.0)
    votes = [p * tr for p, tr in zip(pops, turns)]
    vtot = sum(votes) or 1.0
    rows = [{
        "party": code,
        "popShare": round(pops[k], 2),
        "voteShare": round(votes[k] / vtot * 100, 2),
        "turnout": round(turns[k] * 100, 1),
    } for k, code in enumerate(CODES)]
    write_json(rows, "partyPopulation.json")


def build_turnout_verification():
    """Per-force 2024 turnout verification composition: verifiedVoted / matchedNonvoter /
    unmatched (sum to 100). Shows why the reported-vs-verified gap is dominated by voter-file
    match failure, not over-report. Source: pipeline/add_compare_items.py -> turnout_verification.csv."""
    proc = Path(__file__).parent.parent.parent / "data" / "processed"
    rows = read_csv(proc / "turnout_verification.csv")
    def rec(r):
        return {"party": r["party"],
                "verifiedVoted": round(float(r["verifiedVoted"]), 1),
                "matchedNonvoter": round(float(r["matchedNonvoter"]), 1),
                "unmatched": round(float(r["unmatched"]), 1)}
    national = next(rec(r) for r in rows if r["party"] == "ALL")
    parties = [rec(r) for r in rows if r["party"] != "ALL"]
    write_json({"national": national, "parties": parties}, "turnoutVerification.json")


def build_age_distribution():
    """Per-force 2024 age distribution (weighted percentiles p10/q25/median/q75/p90) for the
    range-bar card. Source: pipeline/add_compare_items.py -> age_distribution.csv."""
    proc = Path(__file__).parent.parent.parent / "data" / "processed"
    rows = read_csv(proc / "age_distribution.csv")
    def rec(r):
        return {"party": r["party"], **{k: round(float(r[k]), 1)
                for k in ("p10", "q25", "median", "q75", "p90")}}
    national = next(rec(r) for r in rows if r["party"] == "ALL")
    parties = [rec(r) for r in rows if r["party"] != "ALL"]
    write_json({"national": national, "parties": parties}, "ageDistribution.json")


def build_turnout_lambda_scenario():
    """Gap-compression sweep. λ=0 (Turnout) already emitted by build_turnout_scenario;
    this emits the intermediate stops *TurnoutL10/L20/L30.json — each force's turnout gap
    closed 10/20/30% toward parity (10 = plausible ceiling, 20–30 = stress). All-parties
    path only; every office + primary responds so the whole app tracks the slider."""
    for l in TURNOUT_STOPS:
        d = OUTPUTS / f"pure_multi_turnout_l{l}"
        _build_turnout_variant(d, f"TurnoutL{l}")
        build_pure_multi_primary(d, f"pureMultiPrimaryTurnoutL{l}.json")
        build_pure_multi_primary_buckets(d, f"pureMultiPrimaryBucketsTurnoutL{l}.json")
        build_pure_multi_primary_state_shares(d, f"pureMultiPrimaryStageSharesTurnoutL{l}.json")


def build_turnout_crossover_triple():
    """Turnout compression outside the RawMulti+double path: the Crossover (FD) senate
    and house, plus the triple-Wyoming house for both pipelines — so the slider works
    on every scenario × Wyoming combination. λ=0 → 'Turnout', 10/20/30 → 'TurnoutLNN'."""
    for l, sfx in [(0, "Turnout"), (10, "TurnoutL10"), (20, "TurnoutL20"), (30, "TurnoutL30")]:
        suf = "_turnout" if l == 0 else f"_turnout_l{l}"
        fd  = OUTPUTS / ("factor_deviation" + suf)
        fdt = OUTPUTS / ("factor_deviation_triple" + suf)
        pmt = OUTPUTS / ("pure_multi_triple" + suf)
        build_fd_senate(fd, f"fdSenateCondorcet{sfx}.json", f"fdSenateIRV{sfx}.json")
        build_fd_house_seats(fd, f"fdHouseSeats{sfx}.json")
        build_fd_district_stv_results(fd, f"fdDistrictStvResults{sfx}.json")
        build_house_seats_triple(pmt, f"houseSeatsTriple{sfx}.json")
        build_district_stv_results_triple(pmt, f"districtStvResultsTriple{sfx}.json")
        build_fd_house_seats_triple(fdt, f"fdHouseSeatsTriple{sfx}.json")
        build_fd_district_stv_results_triple(fdt, f"fdDistrictStvResultsTriple{sfx}.json")


if __name__ == "__main__":
    print("Preparing data (native 10-party incl. OAO)...")

    def _run(fn, *args, **kw):
        """Run a builder; skip (keep committed JSON) if its inputs are missing.
        Some intermediate inputs are pruned from this checkout — those builders
        are skipped and their committed output is left in place."""
        try:
            fn(*args, **kw)
        except FileNotFoundError as e:
            print(f"  SKIP {fn.__name__}: missing input ({e.filename})")
        except Exception as e:
            print(f"  SKIP {fn.__name__}: {type(e).__name__}: {e}")

    # Vote models: the native chamber-vote-model chain (results/vote_model.csv,
    # blend_stats.csv, pure_only/senate) is pruned from this checkout, so we
    # keep the committed vote-model JSON and recompute only the Raw-Multi columns
    # from the native pure_multi run (which now includes OAO).
    _run(build_senate_vote_model_wfp, PURE_MULTI_DIR, out_name="senateVoteModel.json")
    _run(build_house_vote_model_wfp,  PURE_MULTI_DIR, out_name="houseVoteModel.json")

    for fn in (
        build_house_seats, build_house_transfers, build_fd_variant_attraction,
        build_house_state_map, build_coalition_profiles, build_transfer_matrix,
        build_cluster_profiles,
        build_fd_senate, build_fd_house_seats, build_fd_primary,
        build_fd_primary_state_winners, build_fd_primary_sankey,
        build_pure_multi_primary_state_winners, build_fd_presidential_election,
        build_fd_profiles, build_pure_multi_primary, build_pure_multi_primary_sankey,
        build_pure_multi_primary_buckets, build_pure_multi_senate,
        build_senate_buckets, build_senate_condorcet,
        build_raw_multi_presidential_election, build_house_seats_gauss,
        build_fptp_disproportionality, build_district_stv_results,
        build_fd_district_stv_results, build_district_county_map,
        build_county_tiers, build_rcv_results,
        build_house_seats_triple, build_fd_house_seats_triple,
        build_house_state_map_triple, build_district_stv_results_triple,
        build_fd_district_stv_results_triple, build_district_county_map_triple,
        build_nosty_scenario,
        build_turnout_scenario,
        build_turnout_lambda_scenario,
        build_turnout_crossover_triple,
        build_party_population,
        build_turnout_verification,
        build_age_distribution,
    ):
        _run(fn)
    print("Done.")
